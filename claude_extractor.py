"""
Claude-assisted extraction — the PRIMARY parsing path for this project.

Every vendor sends a completely different packing-slip layout (confirmed by
Paula), so writing one deterministic parser per vendor doesn't scale — each new
vendor would block on engineering work. This module reads a spreadsheet's raw
cell grid and asks Claude to return the shipment lines as validated structured
JSON. `parse_packing_slip.py` remains a free fast path for Inprotex's known
format only; see architecture doc section 4.1.

Design decisions worth knowing:

  - **Cell grid, not a screenshot.** The sheet is read with openpyxl and rendered
    as a text grid with real spreadsheet coordinates, so the model can cite
    `PACKING!R42` and a human can find the row. Sending an image would lose the
    exact values and cost far more tokens.
  - **Structured outputs, not prompt-and-parse.** `client.beta.messages.parse()`
    constrains the response to the Pydantic schema in `extraction_schema.py`, so
    a malformed shape is impossible rather than something to defend against.
  - **Per-sheet calls, and row windows for oversized sheets.** Truncating a grid
    to fit a token budget would silently drop lines. Instead each worksheet is
    its own call, and a sheet too large for one call is split into windows that
    repeat the sheet's header rows for context — with a loud warning, because a
    split can separate a PO/style block header from its rows.
  - **Nothing fails silently.** A truncated response (`stop_reason ==
    "max_tokens"`) raises rather than returning partial lines. A refusal raises.
    Rows the model can't read confidently come back flagged low-confidence, and
    anything that looked like line data but couldn't be parsed lands in
    `unparsed_regions`. All of it routes to human review.

Requires an Anthropic API key: set ANTHROPIC_API_KEY, or run `ant auth login`
(the SDK picks up the stored profile with no env var). Not needed for the
Inprotex fast path or for the offline tests.
"""

from __future__ import annotations

import base64
import logging
import os
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional, Sequence, Union

from extraction_schema import (
    PackingSlipExtraction,
    ShippingAdviceExtraction,
)

logger = logging.getLogger(__name__)

#: Secrets live outside the project folder, which is OneDrive-synced — an API key
#: committed to that sync would be uploaded to the cloud and inherited by anyone
#: the folder is shared with. Same reasoning as the NetSuite private key.
SECRETS_ENV_PATH = Path.home() / ".po-agent" / ".env"


def _load_secrets_env(path: Path = SECRETS_ENV_PATH) -> bool:
    """
    Load ANTHROPIC_API_KEY (and anything else) from ~/.po-agent/.env.

    Runs on import so `anthropic.Anthropic()` finds the key without every caller
    having to remember to export it. Deliberately does NOT override variables
    already set in the environment — an explicit `set ANTHROPIC_API_KEY=...` for
    a one-off run should win over the file. Silent no-op if the file is absent
    or python-dotenv isn't installed; `credentials_available()` is what reports
    a missing credential, and it does so at the point of use.
    """
    if not path.is_file():
        return False
    try:
        from dotenv import load_dotenv
    except ImportError:  # pragma: no cover -- dotenv is a declared dependency
        logger.debug("python-dotenv not installed; not loading %s", path)
        return False
    loaded = load_dotenv(path, override=False)
    if loaded:
        logger.debug("Loaded secrets from %s", path)
    return bool(loaded)


_load_secrets_env()

# Claude Opus 5. Effort defaults to `high`, which is what this task wants —
# reading an unfamiliar spreadsheet layout is exactly the judgment call worth
# spending reasoning on, and a misread line costs a human review cycle.
DEFAULT_MODEL = "claude-opus-5"

#: Output cap per call. Kept under the SDK's non-streaming timeout guard; the
#: chunking below keeps each call's output well inside it. A response that hits
#: this raises instead of returning partial lines.
DEFAULT_MAX_TOKENS = 16000

#: Generous HTTP timeout — a large grid at high effort can legitimately take
#: minutes, and a client-side timeout mid-extraction is indistinguishable from
#: a failure.
DEFAULT_TIMEOUT_SECONDS = 900

#: Chunking budget per API call. Whichever limit is reached first wins.
DEFAULT_MAX_ROWS_PER_CALL = 400
DEFAULT_MAX_CHARS_PER_CALL = 120_000

#: Rows repeated at the top of every window of a split sheet, so a window that
#: starts mid-table still carries the column headers.
CONTEXT_HEADER_ROWS = 12

#: Opt into server-side refusal fallbacks. Claude Opus 5 runs safety classifiers
#: that can decline a request; without this a decline just stops. Vendor packing
#: slips are not a domain those classifiers target, so this is belt-and-braces —
#: and it is dropped automatically if the account doesn't have the beta enabled
#: (see `_parse_with_retry`).
FALLBACK_BETA = "server-side-fallback-2026-07-01"


class ExtractionError(Exception):
    """Extraction failed in a way that must not be treated as 'no lines found'."""


class ExtractionTruncated(ExtractionError):
    """The model hit max_tokens — the result is partial and must not be used."""


class ExtractionRefused(ExtractionError):
    """Safety classifiers declined the request."""


class NoPackingSheetFound(ExtractionError):
    """
    A multi-sheet workbook contained no sheet classifiable as a packing list.

    Raised rather than returning zero lines: a silent empty result is
    indistinguishable from "this shipment had nothing in it", which is the worst
    available failure mode. The message names the workbook and every sheet with
    its predicted type, so the decision stays auditable without re-running
    anything.
    """

    def __init__(self, path: Union[str, Path], verdicts: Sequence[Any]):
        self.path = Path(path)
        self.verdicts = list(verdicts)
        listing = "\n".join(
            f"    - '{v.label}': {v.doc_type.value}"
            + (f" -- {v.skip_reason}" if v.skip_reason else "")
            + (f" ({v.reason})" if v.reason else "")
            for v in verdicts
        )
        super().__init__(
            f"{self.path.name}: no sheet classifies as a packing list with per-size "
            f"quantities, so nothing was extracted. Sheets examined:\n{listing}\n"
            "This shipment needs manual attention -- it is NOT an empty shipment."
        )


class DocumentUnreadable(ExtractionError):
    """
    The file could not be opened at all — corrupt, truncated, or encrypted.

    Distinct from "opened fine but the content was unusable", which is what the
    low-confidence flagging handles. This is the failure mode where a vendor's
    attachment arrives half-transferred or password-protected: one bad attachment
    must flag itself and let the rest of the batch continue, not take down the job.
    """

    def __init__(self, path: Union[str, Path], reason: str, original: Optional[BaseException] = None):
        self.path = Path(path)
        self.reason = reason
        self.original = original
        super().__init__(f"{self.path.name}: {reason}")


def _describe_open_failure(exc: BaseException, path: Path) -> str:
    """
    Turn a file-open exception into a reason a human can act on.

    Matches on both exception type and message text, because the libraries are
    inconsistent: a password-protected PDF surfaces from pdfplumber as a generic
    `PdfminerException` whose message is the only thing identifying it as an
    encryption problem.
    """
    import zipfile

    name = type(exc).__name__
    text = str(exc).lower()

    if isinstance(exc, FileNotFoundError):
        return "file not found"
    if isinstance(exc, PermissionError):
        return "permission denied by the operating system (file locked or in use?)"
    if isinstance(exc, zipfile.BadZipFile) or "not a zip file" in text or "bad zip" in text:
        return (
            "not a readable .xlsx — an .xlsx is a zip archive and this one is corrupt or "
            "truncated (a partially transferred attachment looks exactly like this)"
        )
    if "password" in text or "encrypt" in text or "PDFPasswordIncorrect" in name:
        return "encrypted or password-protected — no password is configured, so it cannot be read"
    if "InvalidFileException" in name:
        return "not a format openpyxl can read (a legacy .xls or a renamed file?)"
    if "PDFSyntaxError" in name or "syntax" in text:
        return "corrupt or truncated PDF (invalid structure)"
    if "PdfminerException" in name or "PSException" in name or "PDFException" in name:
        # pdfplumber wraps encryption failures in this, so say both possibilities
        # rather than guessing one and being confidently wrong.
        return f"could not be parsed as a PDF — corrupt, truncated, or encrypted ({name}: {exc})"
    if path.stat().st_size == 0 if path.exists() else False:
        return "file is empty (0 bytes)"
    return f"could not be opened ({name}: {exc})"


def open_workbook(path: Union[str, Path], **kwargs: Any) -> Any:
    """
    openpyxl.load_workbook with explicit handling for unopenable files.

    Raises DocumentUnreadable rather than letting a zipfile/openpyxl exception
    escape into the middle of a batch job.
    """
    import openpyxl

    file_path = Path(path)
    try:
        return openpyxl.load_workbook(file_path, **kwargs)
    except DocumentUnreadable:
        raise
    except Exception as exc:  # noqa: BLE001 -- deliberately broad; classified below
        raise DocumentUnreadable(file_path, _describe_open_failure(exc, file_path), exc) from exc


@contextmanager
def open_pdf(path: Union[str, Path]):
    """
    pdfplumber.open as a context manager, with explicit handling for unopenable
    files (corrupt, truncated, encrypted/password-protected).
    """
    import pdfplumber

    file_path = Path(path)
    pdf = None
    try:
        pdf = pdfplumber.open(str(file_path))
    except Exception as exc:  # noqa: BLE001 -- deliberately broad; classified below
        raise DocumentUnreadable(file_path, _describe_open_failure(exc, file_path), exc) from exc
    try:
        yield pdf
    finally:
        try:
            pdf.close()
        except Exception:  # noqa: BLE001 -- a close failure must not mask the real result
            logger.debug("Ignoring error while closing %s", file_path.name)


# ---------------------------------------------------------------------------
# Grid rendering
# ---------------------------------------------------------------------------


def _col_letter(idx: int) -> str:
    """1 -> A, 27 -> AA."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _fmt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\t", " ").strip()


@dataclass
class SheetGrid:
    """A worksheet materialized as a rectangular grid of formatted strings."""

    name: str
    rows: list[list[str]]  # rows[0] is spreadsheet row 1
    first_col: int  # 1-indexed spreadsheet column of rows[*][0]

    @property
    def n_rows(self) -> int:
        return len(self.rows)

    @property
    def is_empty(self) -> bool:
        return not any(any(cell for cell in row) for row in self.rows)

    def render(self, start_row: int = 1, end_row: Optional[int] = None) -> str:
        """
        Render rows [start_row, end_row] (1-indexed, inclusive) as a text grid.

        Real spreadsheet row numbers and column letters are preserved so the
        model's `source_hint` values point at cells a human can actually open.
        """
        end_row = self.n_rows if end_row is None else min(end_row, self.n_rows)
        width = max((len(r) for r in self.rows), default=0)
        header = "row | " + " | ".join(_col_letter(self.first_col + i) for i in range(width))
        out = [header, "-" * min(len(header), 200)]
        for r in range(start_row, end_row + 1):
            row = self.rows[r - 1]
            if not any(cell for cell in row):
                continue  # skip blank rows -- they carry no information
            padded = row + [""] * (width - len(row))
            out.append(f"{r:>3} | " + " | ".join(padded))
        return "\n".join(out)


def read_workbook_grids(xlsx_path: Union[str, Path]) -> list[SheetGrid]:
    """
    Read every worksheet into a SheetGrid, trimming empty edge rows/columns.

    `data_only=True` so formula cells yield their cached values — the same
    setting `parse_packing_slip.py` uses.
    """
    wb = open_workbook(xlsx_path, data_only=True)
    grids: list[SheetGrid] = []
    for name in wb.sheetnames:
        ws = wb[name]
        raw = [[_fmt(c) for c in row] for row in ws.iter_rows(values_only=True)]
        if not raw:
            grids.append(SheetGrid(name=name, rows=[], first_col=1))
            continue

        # Trim trailing empty columns; find the first non-empty column so a
        # sheet whose data starts at column F doesn't waste tokens on A-E.
        width = max(len(r) for r in raw)
        raw = [r + [""] * (width - len(r)) for r in raw]
        non_empty_cols = [c for c in range(width) if any(r[c] for r in raw)]
        if not non_empty_cols:
            grids.append(SheetGrid(name=name, rows=[], first_col=1))
            continue
        lo, hi = non_empty_cols[0], non_empty_cols[-1]
        trimmed = [r[lo : hi + 1] for r in raw]

        # Trim trailing empty rows only -- leading rows are kept so printed row
        # numbers stay aligned with the real spreadsheet.
        while trimmed and not any(trimmed[-1]):
            trimmed.pop()

        grids.append(SheetGrid(name=name, rows=trimmed, first_col=lo + 1))
    return grids


def plan_windows(
    grid: SheetGrid,
    max_rows: int = DEFAULT_MAX_ROWS_PER_CALL,
    max_chars: int = DEFAULT_MAX_CHARS_PER_CALL,
) -> list[tuple[int, int]]:
    """
    Split a sheet into (start_row, end_row) windows that each fit one API call.

    Returns a single window when the whole sheet fits — the normal case. Windows
    never overlap in the rows they claim; `_render_window` re-attaches the
    sheet's header rows for context instead, and the caller warns when more than
    one window is needed.
    """
    if grid.n_rows == 0:
        return []
    if grid.n_rows <= max_rows and len(grid.render()) <= max_chars:
        return [(1, grid.n_rows)]

    windows: list[tuple[int, int]] = []
    start = 1
    while start <= grid.n_rows:
        end = min(start + max_rows - 1, grid.n_rows)
        # Shrink until the rendered window fits the char budget.
        while end > start and len(grid.render(start, end)) > max_chars:
            end -= max(1, (end - start) // 4)
        windows.append((start, end))
        start = end + 1
    return windows


#: Approximate character width used to reconstruct PDF column alignment as
#: fixed-width text. Derived per page from the actual glyph metrics; this is only
#: the floor, to stop a page of wide glyphs collapsing into one column.
MIN_PDF_CHAR_WIDTH = 3.0


def _estimate_char_width(words: list[dict]) -> float:
    """Median per-character width across a page's words."""
    widths = [
        (w["x1"] - w["x0"]) / len(w["text"])
        for w in words
        if w.get("text") and (w["x1"] - w["x0"]) > 0
    ]
    if not widths:
        return 5.0
    widths.sort()
    return max(MIN_PDF_CHAR_WIDTH, widths[len(widths) // 2])


def render_pdf_page_layout(page: Any) -> str:
    """
    Render a PDF page as fixed-width text that PRESERVES column alignment.

    `page.extract_text()` flattens a table into space-separated tokens, which
    destroys the one thing a reader needs from a numeric table: which column each
    figure sits under. On a real Symmetry inspection report that ambiguity is
    load-bearing -- the size header runs `XS S M L XL 2XL` but a colour's row
    holds only five figures, and whether they mean XS-XL or S-2XL decides whether
    110 units are M or L. The coordinates say XS-XL (values at x0 225/258/289/
    324/357 align to headers at 225/260/291/326/356, with 2XL at 387 empty), but
    the flattened text cannot express that.

    So each word is placed at a character column derived from its x-position,
    reconstructing the visual table. Blank cells stay blank.
    """
    words = page.extract_words()
    if not words:
        return ""
    char_width = _estimate_char_width(words)

    lines: list[list[dict]] = []
    for word in sorted(words, key=lambda w: (round(w["top"], 1), w["x0"])):
        for line in lines:
            if abs(line[0]["top"] - word["top"]) <= 2.5:
                line.append(word)
                break
        else:
            lines.append([word])

    rendered: list[str] = []
    for line in lines:
        line.sort(key=lambda w: w["x0"])
        buf = ""
        for word in line:
            col = int(round(word["x0"] / char_width))
            if col < len(buf):
                col = len(buf) + 1  # never overwrite; keep at least one space
            buf += " " * (col - len(buf)) + word["text"]
        rendered.append(buf.rstrip())
    return "\n".join(rendered)


def read_pdf_layouts(pdf_path: Union[str, Path]) -> list[tuple[str, str]]:
    """
    Column-preserving text for every page of a PDF.

    Returns [(page_label, rendered_text), ...], skipping pages with no text
    layer (scans, photo pages) -- those are reported by the caller rather than
    silently ignored.
    """
    out: list[tuple[str, str]] = []
    with open_pdf(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            rendered = render_pdf_page_layout(page)
            if rendered.strip():
                out.append((f"page {i}", rendered))
    return out


@dataclass
class SourceDocument:
    """One rendered source document (or page) for a multi-document extraction."""

    label: str  # e.g. "FA26 7TH ... INSPECTION REPORT.pdf (page 1)"
    rendered: str
    kind: str = ""  # "packing list", "inspection report", ... if known


def _render_window(grid: SheetGrid, start: int, end: int, split: bool) -> str:
    body = grid.render(start, end)
    if not split:
        return f"=== SHEET: {grid.name} (rows 1-{grid.n_rows}) ===\n{body}"
    header = grid.render(1, min(CONTEXT_HEADER_ROWS, grid.n_rows))
    return (
        f"=== SHEET: {grid.name} — PARTIAL WINDOW, rows {start}-{end} of {grid.n_rows} ===\n"
        f"[The sheet's first rows are repeated below for column context. Extract ONLY\n"
        f" lines whose data rows fall within rows {start}-{end}; the context header is\n"
        f" for interpreting columns, not for extraction.]\n"
        f"--- context header (rows 1-{min(CONTEXT_HEADER_ROWS, grid.n_rows)}) ---\n{header}\n"
        f"--- window (rows {start}-{end}) ---\n{body}"
    )


# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------

# Stable across every call, so it sits in front of the cache breakpoint.
PACKING_SYSTEM_PROMPT = """\
You extract shipment line items from vendor packing slips for a purchase-order \
reconciliation pipeline at an apparel company. Every vendor uses a different \
spreadsheet layout, so work out the structure from the grid in front of you \
rather than expecting a fixed shape.

You are given one worksheet rendered as a text grid with real spreadsheet row \
numbers and column letters. Return one line per PO / style / colour / size \
combination, in long format — if the sheet has sizes across columns, transpose \
them into one line each.

Rules:

1. Emit values exactly as the vendor printed them. Do not normalize, translate, \
or tidy colour codes or size labels. If the sheet says XXL, emit XXL — a \
downstream component maps vendor size labels to the ERP's own convention, and \
it will be wrong if you have already converted them.

2. Read quantities from the recap/summary block when the sheet has one (a \
per-colour, per-size total table). Do not sum individual carton rows yourself \
when a recap total is printed — and if a recap total disagrees with the carton \
rows, extract the recap figure and add a warning saying so.

3. Skip rows whose quantity is zero or blank.

4. Never invent a value. If a row's PO number, style, colour, size, or quantity \
is not determinable from the sheet, still emit the line, set confidence to \
'low', and say what was ambiguous in the note. A flagged line gets a human's \
attention; a silently guessed line becomes a wrong order quantity in the ERP.

5. Be honest about confidence. 'high' means every field is unambiguous. \
'medium' means you inferred a field from layout or surrounding context. 'low' \
means you are guessing. Do not mark inferences 'high'.

6. If you see something that looks like shipment line data but cannot turn it \
into a line, describe it in unparsed_regions with its location and what blocked \
you. Leaving data out silently is the one unacceptable outcome; reporting that \
you could not read it is always acceptable.

7. Put anything else a reviewer should know in warnings — totals that don't \
reconcile, duplicated rows, ambiguous units, multiple PO blocks that might have \
been confused.

Set source_hint to SHEET!R<row> for the row you read each line from, so a human \
can find it in the original file."""

# Used when one shipment's data is split across several documents. Kept separate
# from PACKING_SYSTEM_PROMPT so each stays a stable, cacheable prefix.
MULTI_DOC_SYSTEM_PROMPT = """\
You extract shipment line items for a purchase-order reconciliation pipeline at \
an apparel company. You are given SEVERAL documents that all relate to the SAME \
shipment, and you must combine them into one set of lines.

This matters because no single document is guaranteed to be complete. A \
customs-style invoice / packing list may give quantities per style and colour \
with no size breakdown at all, while a QC or final-inspection report for the \
same PO carries the per-size table. Neither alone is usable; together they are. \
Read all of them before emitting anything, and take each field from whichever \
document actually states it.

Rules:

1. The documents are rendered as fixed-width text that PRESERVES the original \
column alignment. Column position is meaningful data. Read every figure by the \
column header it sits under, never by counting values left to right. If a row \
has fewer figures than the header has columns, some cell is BLANK — work out \
which from the horizontal alignment, and do not shift the values across. \
Getting this wrong silently assigns units to the wrong size.

2. Extract SHIPPED quantities, not ordered quantities. Inspection reports \
routinely show both (e.g. an "ORDER Qty" row and a "SHIPMENT Qty" row); the \
pipeline is reconciling what actually shipped. If you cannot tell which is \
which, set confidence to 'low' and say so.

3. Cross-check between documents and report the result in warnings. If one \
document gives a colour's total and another breaks that colour into sizes, the \
sizes must sum to the total. State whether they reconciled. A mismatch is \
important information, not a reason to silently prefer one source.

4. Emit values exactly as printed. Do not normalize size labels or colour names \
(2XL stays 2XL; do not convert it) — a downstream component owns that mapping \
and will be wrong if you have already converted.

5. Skip rows whose shipped quantity is zero or blank.

6. Set source_hint to the document label and location you read each line from, \
e.g. 'INSPECTION REPORT page 1, SHIPMENT Qty row, BLACK'. When a line's \
quantity and its size come from different documents, name both.

7. Never invent a value. If a required field (PO, style, colour, size, \
quantity) is not stated in ANY of the documents, still emit the line, set \
confidence to 'low', and explain in the note. In particular, if the documents \
give you colour-level totals but no size breakdown anywhere, say so plainly \
rather than inventing a size split.

8. Put anything else a reviewer should know in warnings, and describe in \
unparsed_regions anything that looked like shipment line data you could not \
turn into a line. Silently omitting data is the one unacceptable outcome."""

SHIPPING_SYSTEM_PROMPT = """\
You extract shipment-level fields from freight documents (shipping advices, \
booking confirmations, arrival notices) for a purchase-order reconciliation \
pipeline. Layouts vary by forwarder.

Extract: vendor invoice number, house air waybill (HAWB), master air waybill \
(MAWB), estimated time of departure (ETD), and estimated time of arrival (ETA) \
at the destination port.

Rules:

1. Emit dates exactly as printed. Do not reformat, convert timezones, or \
compute anything. Do not adjust the ETA for inland transit — a later step \
handles that with a human confirming it.

2. ETD and ETA are easy to swap when a routing table lists two bare dates. If \
you are inferring which is which from column position rather than an explicit \
label, set confidence to 'low' and say so in the note. A swapped ETA silently \
becomes a wrong receipt date in the ERP.

3. Use an empty string for any field the document does not contain. Never \
invent or infer a waybill or invoice number.

4. Put anything else a reviewer should know in warnings."""


# ---------------------------------------------------------------------------
# Extractor
# ---------------------------------------------------------------------------


class ClaudeExtractor:
    """
    Wraps the Anthropic API calls for both document types.

    Inject a client for testing:  ClaudeExtractor(client=FakeClient())
    """

    def __init__(
        self,
        client: Any = None,
        model: str = DEFAULT_MODEL,
        max_tokens: int = DEFAULT_MAX_TOKENS,
        timeout: int = DEFAULT_TIMEOUT_SECONDS,
        use_fallbacks: bool = True,
    ):
        self._client = client
        self.model = model
        self.max_tokens = max_tokens
        self.timeout = timeout
        self.use_fallbacks = use_fallbacks
        self.last_usage: dict[str, int] = {}

    @property
    def client(self) -> Any:
        if self._client is None:
            import anthropic

            # Zero-arg constructor resolves ANTHROPIC_API_KEY, ANTHROPIC_AUTH_TOKEN,
            # or an `ant auth login` profile -- don't hardcode a key.
            self._client = anthropic.Anthropic(timeout=self.timeout)
        return self._client

    # -- core call ----------------------------------------------------------

    def _parse_with_retry(self, *, schema: type, system: list[dict], content: list[dict]) -> Any:
        """
        One structured-output call, with the refusal-fallback beta dropped if the
        account doesn't have it.

        `output_config` is deliberately not passed: `parse()` builds it from
        `output_format`, and effort already defaults to `high`, which is the
        right setting for reading an unfamiliar layout.
        """
        kwargs: dict[str, Any] = {
            "model": self.model,
            "max_tokens": self.max_tokens,
            "system": system,
            "messages": [{"role": "user", "content": content}],
            "output_format": schema,
            "thinking": {"type": "adaptive"},
        }
        if self.use_fallbacks:
            kwargs["betas"] = [FALLBACK_BETA]
            kwargs["fallbacks"] = "default"

        try:
            response = self.client.beta.messages.parse(**kwargs)
        except Exception as exc:  # noqa: BLE001 -- classify, then re-raise
            text = str(exc)
            retryable_beta_problem = self.use_fallbacks and (
                "fallback" in text.lower() or "beta" in text.lower()
            )
            if not retryable_beta_problem:
                raise
            logger.warning(
                "Refusal-fallback beta rejected (%s); retrying without it. "
                "Extraction still works; a safety decline would just surface as an error.",
                text.splitlines()[0][:160],
            )
            self.use_fallbacks = False
            kwargs.pop("betas", None)
            kwargs.pop("fallbacks", None)
            response = self.client.beta.messages.parse(**kwargs)

        stop = getattr(response, "stop_reason", None)
        if stop == "refusal":
            details = getattr(response, "stop_details", None)
            raise ExtractionRefused(
                "Anthropic safety classifiers declined this extraction "
                f"(category: {getattr(details, 'category', None)}). "
                "This is unexpected for a packing slip -- inspect the document before retrying."
            )
        if stop == "max_tokens":
            raise ExtractionTruncated(
                f"Response hit max_tokens ({self.max_tokens}); the extraction is partial and "
                "will NOT be used. Lower max_rows_per_call so each window produces fewer "
                "lines, or raise max_tokens."
            )

        usage = getattr(response, "usage", None)
        if usage is not None:
            for key in ("input_tokens", "output_tokens", "cache_read_input_tokens", "cache_creation_input_tokens"):
                value = getattr(usage, key, None)
                if value:
                    self.last_usage[key] = self.last_usage.get(key, 0) + value

        parsed = getattr(response, "parsed_output", None)
        if parsed is None:
            raise ExtractionError(
                "Structured output came back empty despite a non-refusal stop reason "
                f"({stop!r}). Refusing to treat this as 'no lines found'."
            )
        return parsed

    # -- packing slips ------------------------------------------------------

    def extract_sheet(
        self,
        grid: SheetGrid,
        start_row: int,
        end_row: int,
        split: bool,
        source_file: str = "",
    ) -> PackingSlipExtraction:
        """Extract one window of one worksheet."""
        rendered = _render_window(grid, start_row, end_row, split)
        preamble = f"File: {source_file}\n\n" if source_file else ""
        return self._parse_with_retry(
            schema=PackingSlipExtraction,
            # Cache the stable instructions; the volatile grid goes after, in
            # the user turn, so the cached prefix survives across sheets/files.
            system=[
                {
                    "type": "text",
                    "text": PACKING_SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            content=[{"type": "text", "text": f"{preamble}{rendered}"}],
        )

    def _select_packing_sheets(
        self, path: Path, grids: list[SheetGrid]
    ) -> tuple[list[SheetGrid], list[str]]:
        """
        Choose which worksheets of a workbook actually hold packing-list data.

        **A multi-sheet workbook is not one document, it is N documents in a
        container.** Inprotex's file is the proof: four non-empty sheets, of which
        only `PACKING` holds line data, while `(+DIV CHARGE).`, `SHIPMENT` and
        `INVOICE(PAYMENT)` are invoice views. Extracting all four produced 42
        sizeless lines and a 4x-inflated total (25548 units vs the correct 6387),
        and spent roughly three quarters of the file's input tokens on sheets that
        should never have been read.

        Selection goes through `attachment_classifier.classify_sections`, which is
        the same content-based prompt, schema and decision rule used to classify
        whole attachments -- deliberately not a new heuristic. There is no filename
        to lean on at sheet level, and name-based selection would be wrong anyway:
        Legendz's only sheet is called `PO#1657` and is a perfectly good packing
        list, while a sheet called `PACKING LIST` could be an invoice.

        Classification is skipped when there is only one non-empty sheet: there is
        no selection to make, the file-level triage has already vetted the
        document, and it keeps single-sheet workbooks at zero extra API cost.

        Returns (sheets_to_extract, skip_notes). Raises NoPackingSheetFound if a
        multi-sheet workbook yields no packing-list sheet at all -- silently
        returning zero lines is the worst available outcome here.
        """
        if len(grids) <= 1:
            return grids, []

        from attachment_classifier import classify_sections

        verdicts = classify_sections(
            [(g.name, __import__("attachment_classifier").sheet_preview(g)) for g in grids],
            extractor=self,
        )
        by_name = {v.label: v for v in verdicts}

        keep: list[SheetGrid] = []
        skip_notes: list[str] = []
        for grid in grids:
            verdict = by_name.get(grid.name)
            if verdict is not None and verdict.is_shipment_data:
                keep.append(grid)
                logger.info(
                    "%s: extracting sheet '%s' (%s, sizes=%s)",
                    path.name, grid.name, verdict.doc_type.value, verdict.has_size_breakdown,
                )
                continue
            doc_type = verdict.doc_type.value if verdict else "unclassified"
            reason = verdict.skip_reason if verdict else "no verdict returned"
            detail = verdict.reason if verdict else ""
            note = (
                f"sheet '{grid.name}' NOT extracted — predicted type: {doc_type}; "
                f"reason: {reason}" + (f"; classifier said: {detail}" if detail else "")
            )
            skip_notes.append(note)
            # Logged as well as returned: a wrongly-skipped sheet has to be
            # debuggable months later without re-running the classifier by hand.
            logger.warning("%s: %s", path.name, note)

        if not keep:
            raise NoPackingSheetFound(path, verdicts)

        return keep, skip_notes

    def extract_workbook(
        self,
        xlsx_path: Union[str, Path],
        max_rows_per_call: int = DEFAULT_MAX_ROWS_PER_CALL,
        max_chars_per_call: int = DEFAULT_MAX_CHARS_PER_CALL,
        sheet_names: Optional[Sequence[str]] = None,
    ) -> PackingSlipExtraction:
        """
        Extract every (or selected) worksheet and merge the results.

        Merging is a concatenation, not a dedup: two sheets legitimately holding
        the same PO/style/colour/size is a real condition a human should see, so
        it surfaces as a warning rather than being silently collapsed.
        """
        path = Path(xlsx_path)
        grids = [g for g in read_workbook_grids(path) if not g.is_empty]
        if sheet_names is not None:
            wanted = {n.lower() for n in sheet_names}
            grids = [g for g in grids if g.name.lower() in wanted]
        if not grids:
            raise ExtractionError(
                f"{path.name}: no non-empty worksheets to extract"
                + (f" (looked for {list(sheet_names)})" if sheet_names else "")
            )

        # Extract only the sheets that actually hold packing-list data. An explicit
        # sheet_names filter means the caller already made that decision, so
        # classification is not second-guessed in that case.
        skip_notes: list[str] = []
        if sheet_names is None:
            grids, skip_notes = self._select_packing_sheets(path, grids)

        merged = PackingSlipExtraction(
            vendor_name="", document_summary="", lines=[], unparsed_regions=[], warnings=[]
        )
        merged.warnings.extend(skip_notes)
        summaries: list[str] = []

        for grid in grids:
            windows = plan_windows(grid, max_rows_per_call, max_chars_per_call)
            split = len(windows) > 1
            if split:
                merged.warnings.append(
                    f"Sheet '{grid.name}' ({grid.n_rows} rows) exceeded one call and was read in "
                    f"{len(windows)} windows. A split can separate a PO/style block header from "
                    f"its rows -- verify this sheet's lines carefully."
                )
            for start, end in windows:
                logger.info("Extracting %s!%s-%s%s", grid.name, start, end, " (split)" if split else "")
                part = self.extract_sheet(grid, start, end, split, source_file=path.name)
                merged.lines.extend(part.lines)
                merged.unparsed_regions.extend(part.unparsed_regions)
                merged.warnings.extend(part.warnings)
                if part.vendor_name and not merged.vendor_name:
                    merged.vendor_name = part.vendor_name
                if part.document_summary:
                    summaries.append(f"{grid.name}: {part.document_summary}")

        merged.document_summary = " | ".join(summaries)
        self._flag_duplicates(merged)
        return merged

    def extract_pdf_packing_list(self, pdf_path: Union[str, Path]) -> PackingSlipExtraction:
        """
        Extract a packing list that arrives as a PDF rather than a spreadsheet.

        Symmetry's real packing list is a PDF, so this is a primary path, not an
        edge case. Pages are rendered with `render_pdf_page_layout` so column
        alignment survives -- essential here because a colour that doesn't ship
        in every size leaves blank cells, and reading the figures left-to-right
        by count would assign them to the wrong sizes.

        **All pages go in ONE call.** An earlier version sent each page
        separately, which broke on Symmetry's carton-detail file: the size column
        headers, PO numbers and style/colour labels are printed once on page 1,
        and the carton rows continue onto page 2. Page 2 alone is unattributable
        -- the extractor correctly refused to guess and returned it as unparsed,
        but the fix belongs here, not in the prompt. A multi-page packing list is
        one table split across sheets of paper, and has to be read as one.

        Only if the combined text exceeds the per-call budget are pages grouped,
        and then page 1 is always resent as context, with a warning -- the same
        approach `plan_windows` takes for an oversized worksheet.
        """
        path = Path(pdf_path)
        pages = read_pdf_layouts(path)
        if not pages:
            raise ExtractionError(
                f"{path.name}: no extractable text on any page (likely a scan). "
                "Refusing to report zero lines for a document we could not read."
            )

        groups = self._group_pages(pages)
        merged = PackingSlipExtraction(
            vendor_name="", document_summary="", lines=[], unparsed_regions=[], warnings=[]
        )
        if len(groups) > 1:
            merged.warnings.append(
                f"{path.name} was too large for one call and was read in {len(groups)} groups of "
                f"pages (page 1 resent as context each time). A table continuing across a group "
                f"boundary may be misread -- verify this document's lines carefully."
            )

        summaries: list[str] = []
        for group_index, group in enumerate(groups, 1):
            labels = ", ".join(label for label, _ in group)
            logger.info("Extracting %s: %s", path.name, labels)
            body = "\n\n".join(f"===== {label} =====\n{text}" for label, text in group)
            context = ""
            if len(groups) > 1 and group_index > 1:
                context = (
                    f"[Page 1 is repeated below because it carries the column headers and the "
                    f"PO/style/colour labels that later pages rely on. Extract only lines whose "
                    f"data rows appear in {labels}.]\n"
                    f"===== {pages[0][0]} (context only) =====\n{pages[0][1]}\n\n"
                )
            part = self._parse_with_retry(
                schema=PackingSlipExtraction,
                system=[
                    {
                        "type": "text",
                        "text": PACKING_SYSTEM_PROMPT,
                        "cache_control": {"type": "ephemeral"},
                    }
                ],
                content=[
                    {
                        "type": "text",
                        "text": (
                            f"File: {path.name} — {len(pages)} page(s), all supplied below.\n\n"
                            "This is a PDF rendered as fixed-width text. Column alignment is "
                            "preserved and is meaningful: read every figure by the column header "
                            "above it, never by counting values left to right. Rows routinely "
                            "have fewer figures than there are size columns because a colour "
                            "does not ship in every size.\n\n"
                            "This is ONE table that may continue across pages. Column headers, "
                            "PO numbers and style/colour labels are often printed once on the "
                            "first page and apply to rows on later pages -- carry them forward "
                            "rather than treating each page as a separate document.\n\n"
                            f"{context}{body}"
                        ),
                    }
                ],
            )
            merged.lines.extend(part.lines)
            merged.unparsed_regions.extend(part.unparsed_regions)
            merged.warnings.extend(part.warnings)
            if part.vendor_name and not merged.vendor_name:
                merged.vendor_name = part.vendor_name
            if part.document_summary:
                summaries.append(part.document_summary)

        merged.document_summary = " | ".join(summaries)
        self._flag_duplicates(merged)
        return merged

    @staticmethod
    def _group_pages(
        pages: list[tuple[str, str]], max_chars: int = DEFAULT_MAX_CHARS_PER_CALL
    ) -> list[list[tuple[str, str]]]:
        """
        Group rendered pages into per-call batches, all pages in one batch when
        they fit. Returns [[(label, text), ...], ...].
        """
        total = sum(len(text) for _, text in pages)
        if total <= max_chars:
            return [list(pages)]

        groups: list[list[tuple[str, str]]] = []
        current: list[tuple[str, str]] = []
        size = 0
        for page in pages:
            page_len = len(page[1])
            if current and size + page_len > max_chars:
                groups.append(current)
                current, size = [], 0
            current.append(page)
            size += page_len
        if current:
            groups.append(current)
        return groups

    @staticmethod
    def _flag_duplicates(merged: PackingSlipExtraction) -> None:
        """
        Warn when the same PO/style/colour/size appears more than once.

        Not merged automatically: on a carton-level document the same combination
        legitimately recurs across cartons, and whether those should be summed or
        are a double-read is a judgement for a human, not a silent decision here.
        """
        from canonical import canonical_key

        seen: dict[tuple[str, ...], int] = {}
        for line in merged.lines:
            # Canonical key, so a duplicate is still spotted when two rows differ
            # only by internal whitespace or a dash form.
            key = canonical_key(line.po_number, line.style_number, line.color, line.size)
            seen[key] = seen.get(key, 0) + 1
        dupes = [k for k, n in seen.items() if n > 1]
        if dupes:
            merged.warnings.append(
                f"{len(dupes)} PO/style/colour/size combination(s) appear more than once "
                f"(e.g. {dupes[0]}). Not merged -- a human should decide whether these are "
                f"separate cartons to sum or a double-read."
            )

    def extract_documents(
        self, sources: Sequence[SourceDocument], focus: str = ""
    ) -> PackingSlipExtraction:
        """
        Extract one shipment's lines from SEVERAL complementary documents.

        For shipments where no single document is complete -- e.g. Symmetry sends
        a customs invoice/packing list with quantities only by style and colour,
        plus a final inspection report that happens to carry the per-size table.
        All documents go in one call so the model can cross-reference and
        reconcile them; splitting them into separate calls would make that
        impossible by construction.

        `focus` narrows the extraction (e.g. "PO 1721, style W600001") when the
        documents cover more POs than the shipment being processed.
        """
        if not sources:
            raise ExtractionError("extract_documents() needs at least one source document")

        content: list[dict] = []
        if focus:
            content.append(
                {
                    "type": "text",
                    "text": (
                        f"Extract ONLY: {focus}\n\n"
                        "These documents may cover other purchase orders or styles as well. "
                        "Ignore anything outside the scope above -- but if you see line data "
                        "for it that you are deliberately skipping, note that in warnings so "
                        "nobody thinks it was missed."
                    ),
                }
            )

        manifest = "\n".join(
            f"  {i}. {s.label}" + (f" — {s.kind}" if s.kind else "")
            for i, s in enumerate(sources, 1)
        )
        content.append(
            {
                "type": "text",
                "text": f"{len(sources)} document(s) for this one shipment:\n{manifest}",
            }
        )
        for i, source in enumerate(sources, 1):
            content.append(
                {
                    "type": "text",
                    "text": (
                        f"\n===== DOCUMENT {i}: {source.label}"
                        + (f" ({source.kind})" if source.kind else "")
                        + " =====\n"
                        + source.rendered
                    ),
                }
            )

        return self._parse_with_retry(
            schema=PackingSlipExtraction,
            system=[
                {
                    "type": "text",
                    "text": MULTI_DOC_SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            content=content,
        )

    # -- shipping advice ----------------------------------------------------

    def extract_shipping_advice_text(self, text: str, source_file: str = "") -> ShippingAdviceExtraction:
        preamble = f"File: {source_file}\n\n" if source_file else ""
        return self._parse_with_retry(
            schema=ShippingAdviceExtraction,
            system=[
                {
                    "type": "text",
                    "text": SHIPPING_SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            content=[{"type": "text", "text": f"{preamble}--- document text ---\n{text}"}],
        )

    def extract_shipping_advice_pdf(
        self, pdf_bytes: bytes, source_file: str = ""
    ) -> ShippingAdviceExtraction:
        """
        Send the PDF itself, for scanned documents where text extraction yields
        nothing. More expensive than the text path, so it is a fallback.
        """
        return self._parse_with_retry(
            schema=ShippingAdviceExtraction,
            system=[
                {
                    "type": "text",
                    "text": SHIPPING_SYSTEM_PROMPT,
                    "cache_control": {"type": "ephemeral"},
                }
            ],
            content=[
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": base64.standard_b64encode(pdf_bytes).decode("ascii"),
                    },
                },
                {
                    "type": "text",
                    "text": (
                        f"File: {source_file}\n\nText extraction returned nothing usable for this "
                        "document (it is likely a scan). Read the fields from the document itself."
                    ),
                },
            ],
        )


def credentials_available() -> bool:
    """
    Whether an Anthropic credential is resolvable without an interactive login.

    Checks, in the order the SDK resolves them: ANTHROPIC_API_KEY (which
    `_load_secrets_env` has already populated from ~/.po-agent/.env if that file
    holds one), ANTHROPIC_AUTH_TOKEN, then an `ant auth login` profile. An empty
    value in the .env file counts as absent -- a blank placeholder is not a
    credential.
    """
    if os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_AUTH_TOKEN"):
        return True
    config_dir = Path(os.environ.get("ANTHROPIC_CONFIG_DIR", Path.home() / ".config" / "anthropic"))
    return (config_dir / "credentials").is_dir()
