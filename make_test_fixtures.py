"""
Generate the synthetic test fixtures, so they are reproducible rather than
opaque binaries nobody can regenerate or audit.

Fixtures produced (all written into this folder, all safe to commit):

  fixtures/SD Vendor Payment Request SAMPLE (synthetic).pdf
      Same shape as Symmetry's real payment request — same "Request for
      Payment" headings, same field layout — with an entirely fake bank, account
      number, SWIFT code and recipient name. It exists so the attachment
      classifier's "exclude payment requests" behaviour can be tested without a
      real vendor's banking details sitting in a synced project folder. The real
      file now lives outside the project at
      %USERPROFILE%\\.po-agent\\vendor-documents-private\\ and is not needed by
      any test.

  fixtures/corrupt_truncated.xlsx
      A real xlsx cut off mid-file. An .xlsx is a zip archive, so a truncated one
      raises zipfile.BadZipFile on open — a genuinely unopenable file, which the
      pre-existing malformed-content fixture does not cover.

  fixtures/encrypted_password_protected.pdf
      A PDF encrypted with a password we deliberately do not tell the parser, so
      pdfplumber cannot open it.

  fixtures/multi_sheet_one_packing.xlsx
      Four sheets in the shape of Inprotex's real workbook: an invoice sheet, a
      shipment sheet, a payment sheet, and ONE packing sheet with size columns.
      Guards the sheet-selection fix -- extracting all four is what produced 42
      sizeless lines and a 4x-inflated total on the real file.

  fixtures/no_packing_sheet.xlsx
      Two sheets, neither a packing list. Must raise NoPackingSheetFound rather
      than silently returning zero lines.

Run:
    .venv\\Scripts\\python.exe make_test_fixtures.py

Requires reportlab and pypdf, which are test-fixture-generation dependencies
only — the pipeline itself does not import them.
"""

from __future__ import annotations

import shutil
from pathlib import Path

HERE = Path(__file__).resolve().parent
FIXTURES = HERE / "fixtures"

PAYMENT_REQUEST = FIXTURES / "SD Vendor Payment Request SAMPLE (synthetic).pdf"
CORRUPT_XLSX = FIXTURES / "corrupt_truncated.xlsx"
ENCRYPTED_PDF = FIXTURES / "encrypted_password_protected.pdf"
MULTI_SHEET_XLSX = FIXTURES / "multi_sheet_one_packing.xlsx"
NO_PACKING_XLSX = FIXTURES / "no_packing_sheet.xlsx"

#: Password used to encrypt the fixture. Published on purpose: the point of the
#: fixture is that the *parser* has no password, not that this one is secret.
ENCRYPTED_PDF_PASSWORD = "fixture-password-not-a-secret"

# Entirely invented. Deliberately implausible as real banking coordinates:
# "EXAMPLEBANK" is not a real institution and ZZZZUS00XXX is not a valid SWIFT.
_SYNTHETIC_ROWS = [
    ("Ex Fcty", "01-Jan-26"),
    ("Vendor Name", "SAMPLE VENDOR CO., LTD. (SYNTHETIC)"),
    ("Vendor Invoice", "SAMPLE-INV-0001"),
    ("Total Amount Requested (US$)", "US$1,234.56"),
    ("Purchase Order Number(s)", "9001, 9002"),
    ("Shipping Information", "BY : LCL"),
    ("VESSEL NO.", "SAMPLE VESSEL 0001E"),
    ("ETD:", "01.02.2026"),
    ("ETA:", "14.02.2026"),
]

_SYNTHETIC_BANK = [
    (
        "Primary Receiving Bank",
        "Primary Receiving Bank Swift #",
        "Primary Bank Address",
    ),
    (
        "EXAMPLE BANK (SYNTHETIC)",
        "ZZZZUS00XXX",
        "1 Example Street",
    ),
    ("SAMPLE BRANCH", "", "Sampletown, Nowhere"),
]

_SYNTHETIC_RECIPIENT = [
    (
        "Principal Recipient Account Name",
        "Principal Recipient Account #",
        "Principal Recipient Address",
    ),
    (
        "JANE SAMPLE (SYNTHETIC)",
        "000-000000-000",
        "1 Example Street",
    ),
    ("", "", "Sampletown, Nowhere"),
]

_CHECKLIST = [
    "1 Purchase Order(s) (List Above)",
    "2 Itemized Invoice (In US Dollars) (Provide Copy)",
    "3 Packing List (Provide Copy)",
    "4 Bill of Lading or Shipping Information (Provide Copy)",
    "5 Banking Instructions (List Above)",
]


def make_payment_request(path: Path) -> None:
    """A structurally faithful, entirely fictitious payment request."""
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    y = height - 60

    c.setFont("Helvetica-Bold", 13)
    c.drawString(60, y, "Straight Down Clothing Co.")
    y -= 20
    c.drawString(60, y, "Request for Payment")
    y -= 12
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(60, y, "SYNTHETIC TEST FIXTURE — all names, banks and account numbers below are invented.")
    y -= 24

    c.setFont("Helvetica", 10)
    for label, value in _SYNTHETIC_ROWS:
        c.drawString(60, y, label)
        c.drawString(260, y, value)
        y -= 16

    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.drawString(60, y, "Banking Information - Wire Transfer")
    y -= 18
    c.setFont("Helvetica", 9)
    for row in _SYNTHETIC_BANK:
        for x, cell in zip((60, 240, 400), row):
            if cell:
                c.drawString(x, y, cell)
        y -= 14

    y -= 10
    c.setFont("Helvetica", 9)
    for row in _SYNTHETIC_RECIPIENT:
        for x, cell in zip((60, 240, 400), row):
            if cell:
                c.drawString(x, y, cell)
        y -= 14

    y -= 14
    c.setFont("Helvetica-Bold", 9)
    c.drawString(60, y, "Please include the following items with each payment request")
    y -= 16
    c.setFont("Helvetica", 9)
    for item in _CHECKLIST:
        c.drawString(60, y, item)
        y -= 14

    y -= 10
    c.setFont("Helvetica-Bold", 8)
    c.drawString(60, y, "PLEASE COMPLETE THIS FORM AND SUBMIT COPIES WITH EACH PAYMENT REQUEST")
    c.showPage()
    c.save()


def make_corrupt_xlsx(path: Path, source: Path | None = None) -> None:
    """
    A truncated xlsx — valid zip header, no central directory.

    Built by cutting a real workbook in half rather than writing random bytes, so
    the failure mode is the realistic one: a partially-transferred attachment.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    if source is None:
        import openpyxl

        tmp = path.with_suffix(".tmp.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["PO#9001", 100, "PCS"])
        for i in range(300):
            ws.append([i, "M", 24, "filler to make the file worth truncating"])
        wb.save(tmp)
        source = tmp

    data = source.read_bytes()
    path.write_bytes(data[: len(data) // 2])
    if source.name.endswith(".tmp.xlsx"):
        source.unlink()


def make_multi_sheet_workbook(path: Path) -> None:
    """
    Four sheets shaped like Inprotex's real workbook: three non-packing views and
    exactly one packing sheet carrying size columns.

    Sheet ORDER matters: the packing sheet is deliberately LAST, so a fix that
    merely takes the first sheet would fail this fixture.
    """
    import openpyxl

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    inv = wb.active
    inv.title = "INVOICE(PAYMENT)"
    inv.append(["SAMPLE VENDOR CO., LTD."])
    inv.append(["COMMERCIAL INVOICE"])
    inv.append(["PO#9001", "STYLE#S000001", "COLOR", "BLK", "QTY", 300, "AMOUNT", "US$1,000.00"])
    inv.append(["PO#9001", "STYLE#S000001", "COLOR", "NVY", "QTY", 200, "AMOUNT", "US$700.00"])
    inv.append([None, None, None, None, "TOTAL", 500])

    shp = wb.create_sheet("SHIPMENT")
    shp.append(["SHIPMENT DETAIL"])
    shp.append(["VESSEL", "SAMPLE VESSEL 0001E", "ETD", "2026/02/01", "ETA", "2026/02/14"])
    shp.append(["CTNS", 20, "GROSS WT", 250.5, "NET WT", 240.0, "CBM", 1.25])

    pay = wb.create_sheet("(+DIV CHARGE).")
    pay.append(["DIVIDED CHARGES"])
    pay.append(["PO#9001", "STYLE#S000001", "BLK", 300, "FREIGHT", "US$120.00"])
    pay.append(["PO#9001", "STYLE#S000001", "NVY", 200, "FREIGHT", "US$80.00"])

    pk = wb.create_sheet("PACKING")
    pk.append(["SAMPLE VENDOR CO., LTD."])
    pk.append(["PACKING LIST"])
    pk.append(["PO#9001", 500, "PCS"])
    pk.append(["STYLE#S000001"])
    pk.append(["C/NO.", "COLOR", "S", "M", "L", "TOTAL"])
    pk.append([None, "Q'TY", None, None, None, None])
    pk.append([])
    pk.append([None, "S", "M", "L", "TOTAL"])
    pk.append(["BLK", 100, 120, 80, 300])
    pk.append(["NVY", 60, 90, 50, 200])

    wb.save(path)


def make_no_packing_sheet_workbook(path: Path) -> None:
    """Two sheets, neither of which is a packing list with per-size quantities."""
    import openpyxl

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    inv = wb.active
    inv.title = "INVOICE"
    inv.append(["SAMPLE VENDOR CO., LTD."])
    inv.append(["COMMERCIAL INVOICE"])
    inv.append(["PO#9002", "STYLE#S000002", "COLOR", "BLK", "QTY", 150, "AMOUNT", "US$500.00"])
    inv.append([None, None, None, None, "TOTAL", 150])

    summary = wb.create_sheet("SUMMARY")
    summary.append(["SHIPMENT SUMMARY"])
    summary.append(["TOTAL CARTONS", 10, "TOTAL PCS", 150])
    summary.append(["VESSEL", "SAMPLE VESSEL 0002E"])

    wb.save(path)


def make_encrypted_pdf(path: Path) -> None:
    """A password-protected PDF the parser has no password for."""
    from pypdf import PdfWriter
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    path.parent.mkdir(parents=True, exist_ok=True)
    plain = path.with_suffix(".plain.pdf")
    c = canvas.Canvas(str(plain), pagesize=letter)
    c.setFont("Helvetica", 12)
    c.drawString(60, 700, "SYNTHETIC FIXTURE — password-protected packing list")
    c.drawString(60, 680, "PO#9001  STYLE#S000001  COLOR BLK  S 10  M 20  L 30")
    c.showPage()
    c.save()

    writer = PdfWriter(clone_from=str(plain))
    writer.encrypt(ENCRYPTED_PDF_PASSWORD)
    with open(path, "wb") as fh:
        writer.write(fh)
    plain.unlink()


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    make_payment_request(PAYMENT_REQUEST)
    make_corrupt_xlsx(CORRUPT_XLSX)
    make_encrypted_pdf(ENCRYPTED_PDF)
    make_multi_sheet_workbook(MULTI_SHEET_XLSX)
    make_no_packing_sheet_workbook(NO_PACKING_XLSX)

    print("Wrote fixtures:")
    for path in (PAYMENT_REQUEST, CORRUPT_XLSX, ENCRYPTED_PDF, MULTI_SHEET_XLSX, NO_PACKING_XLSX):
        print(f"  {path.relative_to(HERE)}  ({path.stat().st_size} bytes)")
    print()
    print("The real payment request (with genuine banking details) lives outside")
    print("the project at %USERPROFILE%\\.po-agent\\vendor-documents-private\\ and is")
    print("not referenced by any test.")


if __name__ == "__main__":
    main()
