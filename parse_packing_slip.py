"""
Parser for Inprotex-style vendor packing slips (PACKING tab) + UPS shipping advice PDFs.

Output: a normalized, long-format list of PO line updates:
    PO#, Style#, Color, Size, Quantity, Total Cartons (per PO/style/color), Vendor Invoice No,
    HAWB, ETA Date, Ship (dispatch) Date

This is stage 1 of the PO-update pipeline: turn vendor documents into a clean,
reviewable table. It does NOT touch NetSuite -- that happens in a later stage,
after Paula/Kiko review the proposed changes.

Usage:
    python3 parse_packing_slip.py <packing_slip.xlsx> [<shipping_advice.pdf>] [-o output.csv]
"""

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl

from claude_extractor import open_pdf, open_workbook

SIZE_TOKENS = {"XS", "S", "M", "L", "XL", "2XL", "3XL", "XXL", "XXXL", "OS", "ONE SIZE"}


def parse_packing_sheet(xlsx_path, sheet_name="PACKING"):
    """
    Returns a list of dicts:
      {po_number, style_number, color, size, quantity, cartons_total_qty}
    """
    wb = open_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Sheets available: {wb.sheetnames}")
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    # Materialize the grid (1-indexed rows/cols) for positional lookups.
    grid = []
    for r in range(1, max_row + 1):
        grid.append([ws.cell(row=r, column=c).value for c in range(1, max_col + 1)])

    po_re = re.compile(r"PO#\s*(\S+)")
    style_re = re.compile(r"STYLE#\s*(\S+)")

    results = []
    r = 0
    current_po = None
    current_style = None
    current_block_total_qty = None

    while r < len(grid):
        row = grid[r]
        non_null = [(i, v) for i, v in enumerate(row) if v is not None]

        # 1) Block header: a cell containing "PO#1640" alongside a qty + "PCS"
        po_match = None
        for i, v in non_null:
            if isinstance(v, str):
                m = po_re.search(v)
                if m:
                    po_match = m.group(1)
                    break
        if po_match:
            current_po = po_match
            current_style = None
            # total qty is usually the next numeric value in the row (before "PCS")
            current_block_total_qty = None
            for i, v in non_null:
                if isinstance(v, (int, float)):
                    current_block_total_qty = v
                    break
            r += 1
            continue

        # 2) Style line: "STYLE#W60118"
        style_match = None
        for i, v in non_null:
            if isinstance(v, str):
                m = style_re.search(v)
                if m:
                    style_match = m.group(1)
                    break
        if style_match:
            current_style = style_match
            r += 1
            continue

        # 3) Carton-detail header row: has "C/NO." and "COLOR" and "TOTAL"
        str_vals = {v.strip().upper(): i for i, v in non_null if isinstance(v, str)}
        if "C/NO." in str_vals and "COLOR" in str_vals and "TOTAL" in str_vals:
            color_col = str_vals["COLOR"]
            total_col = str_vals["TOTAL"]
            # size columns are between COLOR and TOTAL
            size_cols = []  # list of (col_index, size_label) -- label filled in once we see it
            for c in range(color_col + 1, total_col):
                size_cols.append(c)

            # The size labels themselves are on this same header row, in those columns
            size_labels_by_col = {}
            for c in size_cols:
                v = row[c] if c < len(row) else None
                if isinstance(v, str) and v.strip():
                    size_labels_by_col[c] = v.strip()

            r += 1
            # Skip the "Q'TY"/"KGS"/"KGS" sub-header row and the size-weight row (0.3, 0.35, ...)
            # then skip individual carton rows until we hit the recap section.
            # Recap section = a row whose non-null string values are exactly the size labels + TOTAL
            recap_header_row_idx = None
            search_r = r
            size_label_set = set(size_labels_by_col.values()) | {"TOTAL"}
            while search_r < len(grid):
                candidate = grid[search_r]
                cand_non_null = [v for v in candidate if v is not None]
                if cand_non_null and all(
                    isinstance(v, str) and v.strip().upper() in size_label_set for v in cand_non_null
                ) and "TOTAL" in [v.strip().upper() for v in cand_non_null if isinstance(v, str)]:
                    recap_header_row_idx = search_r
                    break
                # stop if we hit the next PO# block without finding a recap (defensive)
                if any(isinstance(v, str) and po_re.search(v) for v in candidate if isinstance(v, str)):
                    break
                search_r += 1

            if recap_header_row_idx is not None:
                recap_row = grid[recap_header_row_idx]
                recap_col_to_size = {}
                for c, v in enumerate(recap_row):
                    if isinstance(v, str) and v.strip().upper() != "TOTAL" and v.strip():
                        recap_col_to_size[c] = v.strip()
                recap_total_col = None
                for c, v in enumerate(recap_row):
                    if isinstance(v, str) and v.strip().upper() == "TOTAL":
                        recap_total_col = c

                # Now read subsequent rows: first cell = color code, then per-size qty, then total
                cr = recap_header_row_idx + 1
                while cr < len(grid):
                    crow = grid[cr]
                    crow_non_null = [(i, v) for i, v in enumerate(crow) if v is not None]
                    if not crow_non_null:
                        break
                    first_col, first_val = crow_non_null[0]
                    if not isinstance(first_val, str):
                        break
                    color_code = first_val.strip()
                    # stop condition: hit next PO# block
                    if po_re.search(color_code):
                        break
                    for size_col, size_label in recap_col_to_size.items():
                        qty = crow[size_col] if size_col < len(crow) else None
                        if isinstance(qty, (int, float)) and qty:
                            results.append(
                                {
                                    "po_number": current_po,
                                    "style_number": current_style,
                                    "color": color_code,
                                    "size": size_label,
                                    "quantity": int(qty) if float(qty).is_integer() else qty,
                                }
                            )
                    cr += 1
                r = cr
                continue
            else:
                r = search_r
                continue

        r += 1

    return results


# --- shipping advice parsing -------------------------------------------------
#
# ETD/ETA are resolved by matching each date to the COLUMN its label sits in,
# not by the order the dates appear in the document.
#
# Why: on the real sample (Inprotex/EVA BR012, TPE->LAX) the routing row reads
#   1 BR012 TPE 2026/6/27 19:40 LAX 2026/6/27 16:45
# where the ETA (16:45) is chronologically *before* the ETD (19:40) -- correct
# for an eastbound trans-Pacific flight in local times, but it means "the
# earlier date is the departure" is false. The original implementation took the
# first date as ETD and the second as ETA purely by position, which happened to
# be right for this forwarder and would silently swap them for one that orders
# its columns the other way. A swapped ETA becomes a wrong Expected Receipt Date
# in NetSuite with nothing to indicate it.
#
# The document states the mapping explicitly in its header row:
#   ROUTING INFORMATION PORT OF ORIGN ETD PORT OF DEST ETA
# and the columns are left-aligned to the pixel (ETD label and its date both at
# x0=269.46; ETA label and its date both at x0=462.55). So we anchor on those
# labels' x-positions.
#
# If the labels are absent, or a date can't be attributed to one label clearly
# enough, this returns None for that field rather than guessing -- which routes
# the document to the Claude-assisted extractor (see document_parsers.py), where
# an inferred mapping comes back flagged low-confidence for human review.

DATE_TOKEN_RE = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
TIME_TOKEN_RE = re.compile(r"^\d{1,2}:\d{2}$")

#: Header labels we anchor on, normalized -> output field.
DATE_COLUMN_LABELS = {"ETD": "etd", "ETA": "eta"}

#: A date must be this many points closer to its best-matching label than to the
#: runner-up, or we treat the attribution as ambiguous and decline to guess.
LABEL_MATCH_MIN_MARGIN = 12.0

#: Max horizontal gap between a date token and a following time token for them to
#: be considered one cell (e.g. "2026/6/27" + "19:40").
SAME_CELL_MAX_GAP = 15.0


def _normalize_label(text):
    return text.strip().strip(":.,").upper()


def group_words_into_lines(words, tolerance=2.5):
    """
    Group pdfplumber words into visual lines by their `top` coordinate.

    Each line is {"top": float, "words": [word, ...]} with words left-to-right.
    """
    lines = []
    for word in sorted(words, key=lambda w: (w["top"], w["x0"])):
        for line in lines:
            if abs(line["top"] - word["top"]) <= tolerance:
                line["words"].append(word)
                break
        else:
            lines.append({"top": word["top"], "words": [word]})
    for line in lines:
        line["words"].sort(key=lambda w: w["x0"])
    return sorted(lines, key=lambda line: line["top"])


def _find_date_cells(line_words):
    """
    Date tokens on one line, each merged with an adjacent time token.

    Returns [{"text": "2026/6/27 19:40", "x0": ..., "x1": ...}, ...] in the
    "YYYY/M/D HH:MM" shape matcher._parse_eta_to_date expects.
    """
    cells = []
    i = 0
    while i < len(line_words):
        word = line_words[i]
        if DATE_TOKEN_RE.match(word["text"].strip()):
            text, x0, x1 = word["text"].strip(), word["x0"], word["x1"]
            nxt = line_words[i + 1] if i + 1 < len(line_words) else None
            if nxt and TIME_TOKEN_RE.match(nxt["text"].strip()) and nxt["x0"] - x1 < SAME_CELL_MAX_GAP:
                text = f"{text} {nxt['text'].strip()}"
                x1 = nxt["x1"]
                i += 1
            cells.append({"text": text, "x0": x0, "x1": x1})
        i += 1
    return cells


def _column_distance(a, b):
    """
    Horizontal distance between two cells, tolerant of alignment style.

    Takes the better of left-edge and centre distance so the match works whether
    the forwarder's table is left-aligned (this vendor) or centre-aligned.
    """
    left = abs(a["x0"] - b["x0"])
    centre = abs((a["x0"] + a["x1"]) / 2 - (b["x0"] + b["x1"]) / 2)
    return min(left, centre)


def assign_dates_by_label(lines):
    """
    Map dates to ETD/ETA by the column their header label occupies.

    `lines` is the output of group_words_into_lines. Returns
    {"etd": str|None, "eta": str|None, "notes": [str, ...]} — `notes` records
    anything a human should know (labels missing, ambiguous attribution,
    multi-leg routing).

    Deliberately never falls back to date order: that is the bug this replaces.
    """
    notes = []

    header = None
    label_cells = {}
    for line in lines:
        found = {}
        for word in line["words"]:
            field = DATE_COLUMN_LABELS.get(_normalize_label(word["text"]))
            if field and field not in found:
                found[field] = {"x0": word["x0"], "x1": word["x1"]}
        if found:
            header, label_cells = line, found
            break

    if not label_cells:
        notes.append(
            "no ETD/ETA column labels found in the document -- declining to guess from date "
            "order (the labels are what make the mapping trustworthy)"
        )
        return {"etd": None, "eta": None, "notes": notes}

    missing = [f for f in DATE_COLUMN_LABELS.values() if f not in label_cells]
    if missing:
        notes.append(f"header row has no {', '.join(sorted(m.upper() for m in missing))} label")

    # Data rows: lines below the header that contain date cells.
    rows = []
    for line in lines:
        if line["top"] <= header["top"]:
            continue
        cells = _find_date_cells(line["words"])
        if cells:
            rows.append(cells)

    if not rows:
        notes.append("found ETD/ETA labels but no dated routing row beneath them")
        return {"etd": None, "eta": None, "notes": notes}

    per_row = []
    for cells in rows:
        assigned = {}
        for cell in cells:
            scored = sorted(
                (_column_distance(cell, box), field) for field, box in label_cells.items()
            )
            best_distance, best_field = scored[0]
            if len(scored) > 1 and scored[1][0] - best_distance < LABEL_MATCH_MIN_MARGIN:
                notes.append(
                    f"date {cell['text']!r} sits between the ETD and ETA columns "
                    f"(within {LABEL_MATCH_MIN_MARGIN}pt of both) -- not attributed"
                )
                continue
            if best_field in assigned:
                notes.append(
                    f"two dates matched the {best_field.upper()} column on one row "
                    f"({assigned[best_field]!r}, {cell['text']!r}) -- kept the first"
                )
                continue
            assigned[best_field] = cell["text"]
        per_row.append(assigned)

    if len(per_row) > 1:
        notes.append(
            f"{len(per_row)} routing legs found -- using the first leg's ETD and the last "
            f"leg's ETA, which is the through-shipment reading. Verify if that's not intended."
        )

    etd = next((row["etd"] for row in per_row if row.get("etd")), None)
    eta = next((row["eta"] for row in reversed(per_row) if row.get("eta")), None)
    return {"etd": etd, "eta": eta, "notes": notes}


def parse_shipping_advice_pdf(pdf_path):
    """
    Extract HAWB, MAWB, invoice number, ETD and ETA from a shipping advice PDF.

    HAWB/MAWB/invoice come from label-anchored regexes over the text; ETD/ETA
    come from column-position matching against their header labels (see the
    comment block above). Returns None for any field it cannot determine
    confidently, plus a `parse_notes` list — callers should surface those rather
    than dropping them.
    """
    with open_pdf(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        words = []
        for page in pdf.pages:
            words.extend(page.extract_words())

    def find(pattern, default=None):
        m = re.search(pattern, text)
        return m.group(1).strip() if m else default

    dates = assign_dates_by_label(group_words_into_lines(words))

    return {
        "hawb": find(r"HAWB\s*NO\.?\s*([\w\-]+)"),
        "mawb": find(r"MAWB\s*NO\.?\s*([\w\-]+)"),
        "invoice_no": find(r"INV\s*No\s*:\s*([\w\-]+)"),
        "etd": dates["etd"],
        "eta": dates["eta"],
        "parse_notes": dates["notes"],
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path to vendor packing slip .xlsx")
    ap.add_argument("pdf", nargs="?", help="Path to shipping advice PDF (optional)")
    ap.add_argument("-o", "--output", default="proposed_updates.csv", help="Output CSV path")
    ap.add_argument("--sheet", default="PACKING", help="Sheet name containing the packing list")
    args = ap.parse_args()

    lines = parse_packing_sheet(args.xlsx, sheet_name=args.sheet)

    ship_info = {}
    if args.pdf:
        ship_info = parse_shipping_advice_pdf(args.pdf)

    fieldnames = [
        "po_number",
        "style_number",
        "color",
        "size",
        "quantity",
        "hawb",
        "invoice_no",
        "etd",
        "eta",
    ]
    with open(args.output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for line in lines:
            row = dict(line)
            row.update(
                {
                    "hawb": ship_info.get("hawb"),
                    "invoice_no": ship_info.get("invoice_no"),
                    "etd": ship_info.get("etd"),
                    "eta": ship_info.get("eta"),
                }
            )
            writer.writerow(row)

    print(f"Parsed {len(lines)} PO line(s) from {args.xlsx}")
    if ship_info:
        print(f"Shipping advice: {ship_info}")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
