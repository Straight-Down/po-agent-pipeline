"""
Tests for the parsing layer (extraction_schema / claude_extractor / document_parsers).

Three groups:

  A. Offline tests — synthetic fixtures and a mocked Anthropic client. These run
     with no API key and no network, and cover the routing decisions, the
     low-confidence flagging, the failure modes that must raise rather than
     return partial data, and the handoff into matcher.py.

  B. Real-sample tests — run against the real vendor documents in this folder.
     These are the only tests that prove the parsers work on a real document
     rather than on a fixture written to match my own assumptions.

  C. Live Claude tests — the actual API path, opt-in via --live.

Coverage status (updated 2026-08-10):

  - Inprotex — deterministic parser re-verified against the real file, 77 lines.
  - Legendz — Claude path verified live against hand-derived ground truth: sizes
    run as columns, PO and style share one cell separated by a full-width comma,
    and per-block subtotal rows must not be double-counted as line items.
  - Symmetry — Claude path verified live against its REAL packing lists: the
    'Actual Packing Covering' style/colour/size rollup (primary) and the
    'Actual Packing' carton-by-carton detail, which agree with each other
    exactly (25 keys, 1669 units, matching the document's printed G.TOTAL).

CORRECTION (2026-08-11) — an earlier version of this file claimed Symmetry sends
no size breakdown and that a final inspection report had to fill the gap. **Both
halves of that are wrong, not merely superseded.** The document tested back then
was `SD #1720, 1721 INVOICE, PACKING LIST.pdf`, a customs *invoice* whose
quantities stop at style+colour; its filename says "PACKING LIST" but it is not
one. Symmetry's real packing lists carry full size detail. And inspection reports
are now permanently out of scope as a data source (Paula's ruling) regardless of
what they contain. Tests here assert the invoice is *rejected* by the attachment
classifier, which is the behaviour that prevents repeating the mistake.

That is the generalization evidence the build plan's Risks section asked for:
three real vendors, materially different layouts. It is still three, not thirty
— every new vendor is a new layout, so read a green run as "works on what we've
seen", not "works on anything".

Group C costs API tokens and is opt-in. Without --live the Claude path is only
checked against a mock, and the runner says so in its closing summary.

Run:
    .venv\\Scripts\\python.exe test_parsing.py
    .venv\\Scripts\\python.exe test_parsing.py --live    # also hits the real API
"""

from __future__ import annotations

import argparse
import sys
import tempfile
import traceback
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Optional

import claude_extractor as ce
import document_parsers as dp
from extraction_schema import (
    ExtractedLine,
    PackingSlipExtraction,
    ParseResult,
    ShippingAdviceExtraction,
)

HERE = Path(__file__).resolve().parent

# The two real sample documents (CLAUDE.md "Files in this folder").
REAL_XLSX = HERE / "0626建躍空運成衣 (SD-219國外)Invoice_Packing.xlsx"
REAL_PDF = HERE / "Shipping Advice 6128990769 建躍.pdf"

# Second/third real vendors — the generalization evidence the build plan asks for.
LEGENDZ_XLSX = HERE / "Legendz PL0801- 26ctns.xlsx"
# Symmetry's REAL packing lists — the size-level sources.
SYMMETRY_COVERING = HERE / "SD Actual Packing Covering ^N1720^J 1721.pdf"
SYMMETRY_DETAIL = HERE / "SD Actual Packing ^N1720^J 1721.pdf"
# A customs invoice whose filename says "PACKING LIST" but which has no size
# breakdown. Kept as a fixture precisely because mistaking it for the packing
# list is what produced a wrong conclusion; the classifier must reject it.
SYMMETRY_INVOICE = HERE / "SD #1720, 1721 INVOICE, PACKING LIST.pdf"
# Synthetic stand-in. The REAL payment request contains Symmetry's actual bank
# account number and SWIFT code; it now lives outside the project at
# %USERPROFILE%\.po-agentendor-documents-private\ and no test references it.
# This fixture has the same "Request for Payment" structure and field layout
# with invented banking details -- classification behaviour is all the test needs.
SYMMETRY_PAYMENT_REQUEST = HERE / "fixtures" / "SD Vendor Payment Request SAMPLE (synthetic).pdf"
# Permanently excluded as a data source (Paula, 2026-08-11). Used only to assert
# that exclusion, and as a column-alignment rendering fixture.
SYMMETRY_INSPECTION = HERE / "FA26 7TH W600001 PO1721 FINAL INSPECTION REPORT.pdf"
SYMMETRY_INSPECTION_2 = HERE / "FA26 7TH W520005 PO#1721 FINAL INSPECTION REPORT.pdf"

#: Hand-derived from the Legendz sheet: carton rows aggregated per size, with the
#: per-block subtotal rows (11/14/18) and GRAND TOTAL (19) excluded.
LEGENDZ_EXPECTED = {
    ("1657", "M630018", "DFK", "M"): 148,
    ("1657", "M630018", "DFK", "L"): 205,
    ("1657", "M630018", "DFK", "XL"): 188,
    ("1657", "M630018", "DFK", "2XL"): 32,
    ("1657", "M630018", "MLT", "M"): 37,
    ("1657", "M630018", "MLT", "L"): 111,
    ("1657", "M680009", "DKF", "M"): 128,
    ("1657", "M680009", "DKF", "L"): 200,
}

#: Read off the 'Actual Packing Covering' rollup by column position (verified
#: header XS@56 S@64 M@71 L@79 XL@85 2XL@92 3XL@99). Sums to the document's own
#: printed G.TOTAL of 1669 units / 76 cartons.
#:
#: Several rows carry fewer figures than there are size columns because a colour
#: does not ship in every size — M650022 has no XS, W520005 COCONUT stops at L.
#: Reading these left-to-right by count would shift every quantity onto the wrong
#: size, which is the bug the column-position-aware rendering prevents.
SYMMETRY_EXPECTED = {
    ("1720", "M650022", "NEW INDIGO", "S"): 22,
    ("1720", "M650022", "NEW INDIGO", "M"): 157,
    ("1720", "M650022", "NEW INDIGO", "L"): 219,
    ("1720", "M650022", "NEW INDIGO", "XL"): 150,
    ("1720", "M650022", "NEW INDIGO", "2XL"): 65,
    ("1720", "M650022", "NEW INDIGO", "3XL"): 4,
    ("1721", "W520005", "COCONUT", "XS"): 33,
    ("1721", "W520005", "COCONUT", "S"): 79,
    ("1721", "W520005", "COCONUT", "M"): 65,
    ("1721", "W520005", "COCONUT", "L"): 33,
    ("1721", "W520005", "BLACK", "XS"): 17,
    ("1721", "W520005", "BLACK", "S"): 51,
    ("1721", "W520005", "BLACK", "M"): 66,
    ("1721", "W520005", "BLACK", "L"): 52,
    ("1721", "W520005", "BLACK", "XL"): 26,
    ("1721", "W600001", "COCONUT", "XS"): 38,
    ("1721", "W600001", "COCONUT", "S"): 92,
    ("1721", "W600001", "COCONUT", "M"): 104,
    ("1721", "W600001", "COCONUT", "L"): 57,
    ("1721", "W600001", "COCONUT", "XL"): 18,
    ("1721", "W600001", "BLACK", "XS"): 19,
    ("1721", "W600001", "BLACK", "S"): 71,
    ("1721", "W600001", "BLACK", "M"): 110,
    ("1721", "W600001", "BLACK", "L"): 82,
    ("1721", "W600001", "BLACK", "XL"): 39,
}
SYMMETRY_GRAND_TOTAL = 1669

_results: list[tuple[bool, str, str]] = []
_missing_coverage: list[str] = []


def check(ok: bool, name: str, detail: str = "") -> bool:
    _results.append((bool(ok), name, detail))
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f"  --  {detail}" if detail else ""))
    return bool(ok)


def expect_raises(exc_type, fn, name: str) -> None:
    try:
        fn()
    except exc_type as exc:
        check(True, name, f"raised {exc_type.__name__}: {str(exc).splitlines()[0][:70]}")
    except Exception as exc:  # noqa: BLE001
        check(False, name, f"raised {type(exc).__name__} instead of {exc_type.__name__}: {exc}")
    else:
        check(False, name, "did not raise")


def section(title: str) -> None:
    print()
    print(f"--- {title} " + "-" * max(0, 68 - len(title)))


# ---------------------------------------------------------------------------
# Mock Anthropic client
# ---------------------------------------------------------------------------


class FakeParse:
    """Stands in for client.beta.messages.parse. Records calls, replays queued responses."""

    def __init__(self, responses: list[Any], raise_on_first: Optional[Exception] = None):
        self.responses = list(responses)
        self.calls: list[dict] = []
        self.raise_on_first = raise_on_first

    def __call__(self, **kwargs: Any) -> Any:
        self.calls.append(kwargs)
        if self.raise_on_first is not None and len(self.calls) == 1:
            exc, self.raise_on_first = self.raise_on_first, None
            raise exc
        if not self.responses:
            raise AssertionError("FakeParse called more times than it has responses")
        return self.responses.pop(0)


def fake_client(responses: list[Any], raise_on_first: Optional[Exception] = None) -> tuple[Any, FakeParse]:
    parse = FakeParse(responses, raise_on_first)
    client = SimpleNamespace(beta=SimpleNamespace(messages=SimpleNamespace(parse=parse)))
    return client, parse


def response(parsed: Any, stop_reason: str = "end_turn", **usage: int) -> Any:
    return SimpleNamespace(
        parsed_output=parsed,
        stop_reason=stop_reason,
        stop_details=SimpleNamespace(category="cyber") if stop_reason == "refusal" else None,
        usage=SimpleNamespace(**usage) if usage else None,
    )


def line(po="1662", style="M120246", color="TID", size="S", qty=9, conf="high", note="", hint="PACKING!R42"):
    return ExtractedLine(
        po_number=po, style_number=style, color=color, size=size,
        quantity=qty, confidence=conf, note=note, source_hint=hint,
    )


def packing(lines=None, unparsed=None, warnings=None, vendor="Inprotex", summary="recap block"):
    return PackingSlipExtraction(
        vendor_name=vendor,
        document_summary=summary,
        lines=lines if lines is not None else [line()],
        unparsed_regions=unparsed or [],
        warnings=warnings or [],
    )


# ---------------------------------------------------------------------------
# Synthetic fixtures
# ---------------------------------------------------------------------------


def write_inprotex_like(path: Path) -> None:
    """
    A minimal sheet matching the layout `parse_packing_sheet` expects: PO# block
    header, STYLE# line, carton header, then the recap table it actually reads.

    Uses XXL deliberately, so the matcher's SIZE_ALIASES mapping to NetSuite's
    '2X' gets exercised end to end.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PACKING"
    rows = [
        ["PO#1662", 63, "PCS"],
        ["STYLE#M120246"],
        ["C/NO.", "COLOR", "S", "M", "XXL", "TOTAL"],
        [None, "Q'TY", None, None, None, None],
        [],
        [None, "S", "M", "XXL", "TOTAL"],
        ["TID", 9, 50, 4, 63],
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


def write_unknown_vendor(path: Path) -> None:
    """A completely different layout — must route to the Claude extractor."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipment Detail"
    ws.append(["Order Ref", "SKU", "Colourway", "Size Break", "Units Shipped"])
    ws.append(["PO-2044", "W630010", "COC", "XS", 12])
    ws.append(["PO-2044", "W630010", "COC", "S", 30])
    wb.save(path)


def write_malformed(path: Path) -> None:
    """
    Deliberately broken: an Inprotex-looking sheet whose recap quantities are
    unreadable text and whose PO header is missing a number.

    It must NOT come back as clean high-confidence lines.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PACKING"
    rows = [
        ["PO# (see email)", "TBC", "PCS"],
        ["STYLE#"],
        ["C/NO.", "COLOR", "S", "M", "TOTAL"],
        [None, "Q'TY", None, None, None],
        [],
        [None, "S", "M", "TOTAL"],
        ["TID", "approx 9", "fifty", "?"],
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


# ---------------------------------------------------------------------------
# A. Offline tests
# ---------------------------------------------------------------------------


def test_grid_rendering(tmp: Path) -> None:
    section("grid rendering (cell grid, not a screenshot)")
    path = tmp / "inprotex_like.xlsx"
    write_inprotex_like(path)

    grids = ce.read_workbook_grids(path)
    check(len(grids) == 1 and grids[0].name == "PACKING", "reads worksheets", grids[0].name)

    grid = grids[0]
    rendered = grid.render()
    check("row |" in rendered, "renders a coordinate header row")
    check("PO#1662" in rendered, "cell values preserved verbatim")
    check("XXL" in rendered, "vendor size label preserved (not normalized to 2X)")
    check(
        all(f"{n:>3} |" in rendered for n in (1, 2, 3)),
        "real spreadsheet row numbers present (so source_hint is findable)",
    )
    check("\n  5 |" not in rendered, "blank rows omitted to save tokens")

    # Column letters must reflect the real column, not the trimmed offset.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["D2"] = "starts at D"
    offset_path = tmp / "offset.xlsx"
    wb.save(offset_path)
    offset_grid = ce.read_workbook_grids(offset_path)[0]
    check(offset_grid.first_col == 4, "leading empty columns trimmed", f"first_col={offset_grid.first_col}")
    check("row | D" in offset_grid.render(), "column letters stay true to the sheet")

    check(ce._col_letter(1) == "A" and ce._col_letter(27) == "AA", "column letter maths")


def test_windowing(tmp: Path) -> None:
    section("chunking (per-sheet calls, windows instead of truncation)")
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(1, 1001):
        ws.append([f"row{i}", i, "TID", "S"])
    big = tmp / "big.xlsx"
    wb.save(big)
    grid = ce.read_workbook_grids(big)[0]

    single = ce.plan_windows(grid, max_rows=5000, max_chars=10_000_000)
    check(single == [(1, 1000)], "sheet that fits is a single window", str(single))

    windows = ce.plan_windows(grid, max_rows=400, max_chars=10_000_000)
    check(len(windows) == 3, "oversized sheet is split into windows", str(windows))
    check(windows[0][0] == 1 and windows[-1][1] == 1000, "windows span every row -- no truncation")
    covered = [r for start, end in windows for r in range(start, end + 1)]
    check(len(covered) == len(set(covered)) == 1000, "windows are contiguous and non-overlapping")

    char_windows = ce.plan_windows(grid, max_rows=5000, max_chars=8000)
    check(len(char_windows) > 1, "char budget also forces a split", f"{len(char_windows)} windows")

    split_render = ce._render_window(grid, 401, 800, split=True)
    check("PARTIAL WINDOW" in split_render, "split windows are labelled as partial")
    check("context header" in split_render, "split windows repeat header rows for column context")

    empty = ce.SheetGrid(name="blank", rows=[], first_col=1)
    check(ce.plan_windows(empty) == [], "empty sheet produces no calls")


def test_extractor_call_shape(tmp: Path) -> None:
    section("API call shape")
    path = tmp / "inprotex_like.xlsx"
    write_inprotex_like(path)

    client, parse = fake_client([response(packing(), input_tokens=100, output_tokens=50)])
    extractor = ce.ClaudeExtractor(client=client)
    extractor.extract_workbook(path)

    kwargs = parse.calls[0]
    check(kwargs["model"] == "claude-opus-5", "uses Claude Opus 5", kwargs["model"])
    check(kwargs["output_format"] is PackingSlipExtraction, "structured output bound to the schema")
    check(kwargs["thinking"] == {"type": "adaptive"}, "adaptive thinking", str(kwargs["thinking"]))
    check("output_config" not in kwargs, "no output_config (would clash with output_format; effort defaults to high)")
    check(kwargs["fallbacks"] == "default", "refusal fallbacks opted into by default")
    check(kwargs["betas"] == [ce.FALLBACK_BETA], "fallback beta header set", str(kwargs["betas"]))

    system = kwargs["system"][0]
    check(
        system["cache_control"] == {"type": "ephemeral"},
        "stable system prompt is cached (volatile grid goes in the user turn)",
    )
    check(
        "do not turn XXL into 2X" in system["text"].lower().replace("xxl", "XXL").replace("2x", "2X")
        or "XXL" in system["text"],
        "prompt forbids size normalization (matcher.py owns SIZE_ALIASES)",
    )

    user_text = kwargs["messages"][0]["content"][0]["text"]
    check("PO#1662" in user_text, "grid text passed as structured cells")
    check("image" not in str(kwargs["messages"]), "no image block -- cell grid, not a screenshot")
    check(extractor.last_usage.get("input_tokens") == 100, "token usage accumulated", str(extractor.last_usage))


def test_failure_modes_raise(tmp: Path) -> None:
    section("failure modes raise instead of returning partial data")
    path = tmp / "inprotex_like.xlsx"
    write_inprotex_like(path)

    truncated, _ = fake_client([response(packing(), stop_reason="max_tokens")])
    expect_raises(
        ce.ExtractionTruncated,
        lambda: ce.ClaudeExtractor(client=truncated).extract_workbook(path),
        "max_tokens truncation raises (never returns partial lines)",
    )

    refused, _ = fake_client([response(None, stop_reason="refusal")])
    expect_raises(
        ce.ExtractionRefused,
        lambda: ce.ClaudeExtractor(client=refused).extract_workbook(path),
        "safety refusal raises (stop_reason checked before reading content)",
    )

    empty, _ = fake_client([response(None)])
    expect_raises(
        ce.ExtractionError,
        lambda: ce.ClaudeExtractor(client=empty).extract_workbook(path),
        "empty structured output raises, not treated as 'no lines found'",
    )

    blank = tmp / "blank.xlsx"
    import openpyxl

    openpyxl.Workbook().save(blank)
    client, _ = fake_client([])
    expect_raises(
        ce.ExtractionError,
        lambda: ce.ClaudeExtractor(client=client).extract_workbook(blank),
        "workbook with no usable sheets raises",
    )

    # A rejected fallback beta must degrade gracefully, not kill the extraction.
    client, parse = fake_client(
        [response(packing())], raise_on_first=Exception("400 unsupported beta: server-side-fallback")
    )
    extractor = ce.ClaudeExtractor(client=client)
    extractor.extract_workbook(path)
    check(len(parse.calls) == 2, "fallback-beta rejection retries once without it", f"{len(parse.calls)} calls")
    check("fallbacks" not in parse.calls[1], "retry drops the fallbacks parameter")
    check(extractor.use_fallbacks is False, "extractor remembers not to retry the beta")

    # An unrelated error must propagate, not be swallowed by that retry path.
    client, _ = fake_client([response(packing())], raise_on_first=RuntimeError("connection reset"))
    expect_raises(
        RuntimeError,
        lambda: ce.ClaudeExtractor(client=client).extract_workbook(path),
        "unrelated errors propagate (retry is narrowly scoped)",
    )


def test_low_confidence_flagging(tmp: Path) -> None:
    section("low-confidence rows are FLAGGED, never dropped or silently guessed")
    path = tmp / "malformed.xlsx"
    write_malformed(path)

    extraction = packing(
        lines=[
            line(qty=9, conf="high"),
            line(size="M", qty=50, conf="low", note="quantity cell reads 'fifty', not a number"),
            line(po="", size="XXL", qty=0, conf="low", note="PO header says '(see email)'; no number on the sheet"),
        ],
        unparsed=["PACKING!R7 column E: total cell contains '?' -- could not reconcile"],
        warnings=["PO block header has no PO number"],
    )
    client, _ = fake_client([response(extraction)])
    result = dp.parse_packing_slip(path, extractor=ce.ClaudeExtractor(client=client), force="claude")

    check(len(result.lines) == 3, "all 3 lines kept -- unreadable rows are NOT dropped", str(len(result.lines)))
    check(len(result.low_confidence_lines) == 2, "2 flagged low-confidence", str(len(result.low_confidence_lines)))
    check(all(ln["note"] for ln in result.low_confidence_lines), "each flagged line carries a reviewer note")
    check(result.needs_review is True, "needs_review set")
    check(len(result.unparsed_regions) == 1, "unparsed region surfaced, not swallowed")
    check("PO block header has no PO number" in result.warnings, "model warning propagated")
    # Looked up by identity, not by index: output order is now deterministic
    # (sorted by the semantic key), so index-based assertions are meaningless.
    no_po = [l for l in result.lines if not l["po_number"]]
    check(
        len(no_po) == 1 and no_po[0]["confidence"] == "low",
        "an undeterminable PO stays empty + low rather than being invented",
    )

    # needs_review is conservative: each signal alone is enough.
    check(ParseResult(lines=[], parser="x").needs_review, "empty result needs review")
    clean = ParseResult(lines=[{"confidence": "high"}], parser="x")
    check(not clean.needs_review, "fully clean result does not need review")
    check(ParseResult(lines=[{"confidence": "medium"}], parser="x").needs_review, "medium confidence needs review")
    check(
        ParseResult(lines=[{"confidence": "high"}], parser="x", unparsed_regions=["x"]).needs_review,
        "unparsed region alone needs review",
    )
    check(
        ParseResult(lines=[{"confidence": "high"}], parser="x", warnings=["x"]).needs_review,
        "warning alone needs review",
    )
    check("NEEDS REVIEW" in result.review_summary(), "review_summary states the verdict")


def test_routing(tmp: Path) -> None:
    section("routing: known format -> free parser, everything else -> Claude")
    inprotex = tmp / "inprotex_like.xlsx"
    unknown = tmp / "unknown_vendor.xlsx"
    write_inprotex_like(inprotex)
    write_unknown_vendor(unknown)

    matched, reason = dp.looks_like_inprotex(inprotex)
    check(matched, "Inprotex layout detected", reason)
    matched2, reason2 = dp.looks_like_inprotex(unknown)
    check(not matched2, "unknown layout not misdetected as Inprotex", reason2)
    missing, reason3 = dp.looks_like_inprotex(tmp / "nope.xlsx")
    check(not missing, "unreadable file is a routing answer, not a crash", reason3)

    # Known format: free path, no API call at all.
    client, parse = fake_client([])
    result = dp.parse_packing_slip(inprotex, extractor=ce.ClaudeExtractor(client=client))
    check(result.parser == "inprotex-deterministic", "known format used the free parser", result.parser)
    check(len(parse.calls) == 0, "no API call made for the known format")
    check(len(result.lines) == 3, "3 lines from the recap block (S/M/XXL)", str(len(result.lines)))
    check(
        all(ln["confidence"] == "high" for ln in result.lines),
        "deterministic output marked high confidence",
    )
    check({ln["size"] for ln in result.lines} == {"S", "M", "XXL"}, "vendor size labels kept verbatim")
    check(result.lines[0]["po_number"] == "1662", "PO parsed", result.lines[0]["po_number"])
    # Regression: a clean parse must not be flagged for review just because the
    # router recorded which parser it picked. Routing detail is a note, not a warning.
    check(
        not result.needs_review,
        "a clean deterministic parse does NOT need review (routing notes aren't warnings)",
        result.review_summary(),
    )
    check(any("format sniff" in n for n in result.notes), "routing decision recorded in notes")
    check(result.warnings == [], "no warnings on a clean known-format parse", str(result.warnings))

    # Unknown format: Claude path.
    client, parse = fake_client([response(packing(vendor="Unknown Apparel Co"))])
    result = dp.parse_packing_slip(unknown, extractor=ce.ClaudeExtractor(client=client))
    check(result.parser == "claude-assisted", "unknown format routed to Claude", result.parser)
    check(len(parse.calls) == 1, "one API call per sheet")
    check(any("format sniff" in n for n in result.notes), "routing decision recorded in notes")
    check(
        not result.needs_review,
        "a clean Claude-assisted parse does NOT need review either (or the flag would always be True)",
        result.review_summary(),
    )

    # force overrides.
    client, parse = fake_client([response(packing())])
    result = dp.parse_packing_slip(inprotex, extractor=ce.ClaudeExtractor(client=client), force="claude")
    check(result.parser == "claude-assisted" and len(parse.calls) == 1, "force='claude' bypasses the free path")


def test_deterministic_validation(tmp: Path) -> None:
    section("a partly-failing deterministic parse falls back rather than lying")
    malformed = tmp / "malformed.xlsx"
    write_malformed(malformed)

    matched, reason = dp.looks_like_inprotex(malformed)
    check(matched, "malformed file still sniffs as Inprotex (markers present)", reason)

    client, parse = fake_client([response(packing(lines=[line(conf="low", note="unreadable quantities")]))])
    result = dp.parse_packing_slip(malformed, extractor=ce.ClaudeExtractor(client=client))
    check(
        result.parser == "claude-assisted",
        "deterministic output that fails validation is NOT emitted as high-confidence",
        result.parser,
    )
    check(len(parse.calls) == 1, "fell back to Claude")
    check(
        any("validation" in w or "raised" in w for w in result.warnings),
        "fallback reason is a WARNING (a known format that stopped parsing is an anomaly)",
        str([w[:60] for w in result.warnings]),
    )
    check(result.needs_review, "so the file is flagged for review", result.review_summary())

    problems = dp.validate_deterministic_lines(
        [{"po_number": "1662", "style_number": "", "color": "TID", "size": "S", "quantity": 9}]
    )
    check(bool(problems) and "style_number" in problems[0], "validator catches an empty field", str(problems[:1]))
    check(dp.validate_deterministic_lines([]) == ["parser returned no lines"], "validator catches an empty parse")
    check(
        bool(dp.validate_deterministic_lines(
            [{"po_number": "1", "style_number": "s", "color": "c", "size": "S", "quantity": 0}]
        )),
        "validator rejects a zero quantity",
    )
    check(
        dp.validate_deterministic_lines(
            [{"po_number": "1662", "style_number": "M120246", "color": "TID", "size": "S", "quantity": 9}]
        )
        == [],
        "validator passes a complete line",
    )

    expect_raises(
        ce.ExtractionError,
        lambda: dp.parse_packing_slip(malformed, force="deterministic"),
        "force='deterministic' fails loudly rather than silently using Claude",
    )
    expect_raises(
        FileNotFoundError,
        lambda: dp.parse_packing_slip(tmp / "does_not_exist.xlsx"),
        "missing file raises",
    )


def test_shipping_advice_routing(tmp: Path) -> None:
    section("shipping advice: deterministic regex -> Claude text -> Claude PDF")
    fake_pdf = tmp / "advice.pdf"
    fake_pdf.write_bytes(b"%PDF-1.4 not a real pdf")

    complete = {"hawb": "6128990769", "mawb": "160-12345678", "invoice_no": "SD-219",
                "etd": "2026/6/25 18:00", "eta": "2026/6/27 16:45"}

    # 1. Deterministic path when every required field is found.
    import parse_packing_slip as pps

    saved = pps.parse_shipping_advice_pdf
    try:
        pps.parse_shipping_advice_pdf = lambda p: dict(complete)
        info, parser, warnings = dp.parse_shipping_advice(fake_pdf)
        check(parser == "regex-deterministic", "complete regex parse wins (free)", parser)
        check(info["eta"] == "2026/6/27 16:45" and info["confidence"] == "high", "fields returned as printed")

        # 2. Missing a required field -> Claude on the extracted text.
        pps.parse_shipping_advice_pdf = lambda p: {**complete, "eta": None}
        client, parse = fake_client([
            response(ShippingAdviceExtraction(
                invoice_no="SD-219", hawb="6128990769", mawb="", etd="2026/6/25 18:00",
                eta="2026/6/27 16:45", confidence="low",
                note="two bare dates in the routing row; inferred which is ETA from column order",
                warnings=[],
            ))
        ])
        saved_text = dp.extract_pdf_text
        try:
            dp.extract_pdf_text = lambda p: "ROUTING 2026/6/25 18:00 2026/6/27 16:45"
            info, parser, warnings = dp.parse_shipping_advice(fake_pdf, extractor=ce.ClaudeExtractor(client=client))
            check(parser == "claude-assisted-text", "incomplete regex parse falls back to Claude", parser)
            check(any("missing" in w for w in warnings), "records which field was missing")
            check(info["confidence"] == "low", "swapped-date risk surfaced as low confidence")

            # 3. No text layer (scan) -> send the PDF itself.
            client, parse = fake_client([
                response(ShippingAdviceExtraction(
                    invoice_no="", hawb="6128990769", mawb="", etd="", eta="2026/6/27",
                    confidence="medium", note="read from a scan", warnings=[],
                ))
            ])
            dp.extract_pdf_text = lambda p: "   \n  "
            info, parser, warnings = dp.parse_shipping_advice(fake_pdf, extractor=ce.ClaudeExtractor(client=client))
            check(parser == "claude-assisted-pdf", "empty text layer routes to the PDF path", parser)
            content = parse.calls[0]["messages"][0]["content"]
            check(content[0]["type"] == "document", "PDF sent as a document block")
            check(content[0]["source"]["media_type"] == "application/pdf", "correct media type")
            check(any("scan" in w for w in warnings), "warns that the expensive path was used")
        finally:
            dp.extract_pdf_text = saved_text
    finally:
        pps.parse_shipping_advice_pdf = saved


def test_shipping_date_label_anchoring(tmp: Path) -> None:
    section("ETD/ETA anchored to column LABELS, not date order")
    from parse_packing_slip import assign_dates_by_label, group_words_into_lines

    def w(text: str, x0: float, x1: float, top: float) -> dict:
        return {"text": text, "x0": x0, "x1": x1, "top": top}

    def resolve(words: list[dict]) -> dict:
        return assign_dates_by_label(group_words_into_lines(words))

    # Real sample geometry: ETD column at x0=269.46, ETA column at x0=462.55.
    real = [
        w("ROUTING", 35.71, 81.82, 433), w("INFORMATION", 84.60, 154.60, 433),
        w("PORT", 162.18, 189.96, 433), w("OF", 192.74, 206.63, 433), w("ORIGN", 209.41, 242.19, 433),
        w("ETD", 269.46, 289.46, 433),
        w("PORT", 355.27, 383.05, 433), w("OF", 385.83, 399.72, 433), w("DEST", 402.50, 429.17, 433),
        w("ETA", 462.55, 482.00, 433),
        w("1", 44.75, 50.31, 453), w("BR012", 54.91, 85.48, 453), w("TPE", 162.18, 181.63, 453),
        w("2026/6/27", 269.46, 313.94, 453), w("19:40", 316.72, 341.74, 453),
        w("LAX", 355.27, 374.17, 453),
        w("2026/6/27", 462.55, 507.03, 453), w("16:45", 509.81, 534.83, 453),
    ]
    got = resolve(real)
    check(got["etd"] == "2026/6/27 19:40", "real layout: ETD from the ETD column", str(got["etd"]))
    check(got["eta"] == "2026/6/27 16:45", "real layout: ETA from the ETA column", str(got["eta"]))
    check(got["notes"] == [], "no notes needed on a clean label match", str(got["notes"]))
    check(" " in got["eta"] and "/" in got["eta"], "date+time merged in matcher's expected format", got["eta"])

    # THE REGRESSION THIS CHANGE EXISTS FOR: a forwarder that puts ETA first.
    # Same geometry, labels swapped. The FIRST date in document order is now the
    # ETA -- the old date-order logic would have called it the ETD.
    swapped = [
        w("ROUTING", 35.71, 81.82, 433),
        w("ETA", 269.46, 289.46, 433),
        w("ETD", 462.55, 482.00, 433),
        w("1", 44.75, 50.31, 453),
        w("2026/6/27", 269.46, 313.94, 453), w("16:45", 316.72, 341.74, 453),
        w("2026/6/27", 462.55, 507.03, 453), w("19:40", 509.81, 534.83, 453),
    ]
    got = resolve(swapped)
    check(got["eta"] == "2026/6/27 16:45", "swapped columns: ETA still read from the ETA column", str(got["eta"]))
    check(got["etd"] == "2026/6/27 19:40", "swapped columns: ETD still read from the ETD column", str(got["etd"]))
    check(
        got["eta"] != got["etd"] and got["notes"] == [],
        "swapped columns handled silently and correctly -- date order is irrelevant now",
    )

    # No labels -> refuse to guess (the old code would have guessed from order).
    unlabelled = [
        w("ROUTING", 35.71, 81.82, 433),
        w("1", 44.75, 50.31, 453),
        w("2026/6/27", 269.46, 313.94, 453), w("19:40", 316.72, 341.74, 453),
        w("2026/6/27", 462.55, 507.03, 453), w("16:45", 509.81, 534.83, 453),
    ]
    got = resolve(unlabelled)
    check(got["etd"] is None and got["eta"] is None, "no labels -> no dates returned, not a guess")
    check(any("declining to guess" in n for n in got["notes"]), "explains why it declined", str(got["notes"])[:80])

    # Only one of the two labels present.
    partial = [
        w("ETA", 462.55, 482.00, 433),
        w("2026/6/27", 462.55, 507.03, 453), w("16:45", 509.81, 534.83, 453),
    ]
    got = resolve(partial)
    check(got["eta"] == "2026/6/27 16:45" and got["etd"] is None, "one label present -> only that field filled")
    check(any("no ETD label" in n for n in got["notes"]), "missing label reported", str(got["notes"])[:80])

    # Labels but no dated row beneath.
    got = resolve([w("ETD", 269.46, 289.46, 433), w("ETA", 462.55, 482.00, 433)])
    check(got["etd"] is None and any("no dated routing row" in n for n in got["notes"]), "labels but no data row reported")

    # A date sitting equidistant between both columns must NOT be attributed.
    ambiguous = [
        w("ETD", 100, 120, 400), w("ETA", 130, 150, 400),
        w("2026/6/27", 115, 135, 420),
    ]
    got = resolve(ambiguous)
    check(got["etd"] is None and got["eta"] is None, "ambiguous column position -> not attributed")
    check(any("sits between" in n for n in got["notes"]), "ambiguity explained", str(got["notes"])[:80])

    # Date with no time still works (not every forwarder prints a time).
    got = resolve([
        w("ETD", 269.46, 289.46, 433), w("ETA", 462.55, 482.00, 433),
        w("2026/6/27", 269.46, 313.94, 453), w("2026/6/28", 462.55, 507.03, 453),
    ])
    check(got["etd"] == "2026/6/27" and got["eta"] == "2026/6/28", "date-only cells supported", f"{got['etd']} / {got['eta']}")

    # Multi-leg routing: first leg's ETD, last leg's ETA, and say so.
    multi = [
        w("ETD", 269.46, 289.46, 433), w("ETA", 462.55, 482.00, 433),
        w("2026/6/27", 269.46, 313.94, 453), w("2026/6/28", 462.55, 507.03, 453),
        w("2026/6/29", 269.46, 313.94, 470), w("2026/6/30", 462.55, 507.03, 470),
    ]
    got = resolve(multi)
    check(got["etd"] == "2026/6/27", "multi-leg: first leg's ETD", str(got["etd"]))
    check(got["eta"] == "2026/6/30", "multi-leg: last leg's ETA (through-shipment reading)", str(got["eta"]))
    check(any("routing legs" in n for n in got["notes"]), "multi-leg flagged for verification")

    # A time token far from its date must not be glued on.
    got = resolve([
        w("ETD", 269.46, 289.46, 433), w("ETA", 462.55, 482.00, 433),
        w("2026/6/27", 269.46, 313.94, 453), w("19:40", 420.00, 445.00, 453),
        w("2026/6/28", 462.55, 507.03, 453),
    ])
    check(got["etd"] == "2026/6/27", "distant time token not merged into the date cell", str(got["etd"]))

    # parse_notes must reach the caller as warnings, not be dropped.
    fake_pdf = tmp / "notes.pdf"
    fake_pdf.write_bytes(b"%PDF-1.4")
    import parse_packing_slip as pps

    saved = pps.parse_shipping_advice_pdf
    try:
        pps.parse_shipping_advice_pdf = lambda p: {
            "hawb": "H1", "mawb": "M1", "invoice_no": "I1",
            "etd": "2026/6/27 19:40", "eta": "2026/6/27 16:45",
            "parse_notes": ["2 routing legs found -- using the first leg's ETD"],
        }
        info, parser, warns = dp.parse_shipping_advice(fake_pdf)
        check(parser == "regex-deterministic", "complete label-anchored parse still wins", parser)
        check(any("routing legs" in w_ for w_ in warns), "parse_notes surfaced as a warning", str(warns))
        check("parse_notes" not in info, "parse_notes consumed, not leaked into ship_info", str(list(info)))
    finally:
        pps.parse_shipping_advice_pdf = saved


def test_pdf_layout_rendering(tmp: Path) -> None:
    section("PDF rendering preserves column alignment (the XS-vs-2XL trap)")
    if not SYMMETRY_INSPECTION.exists():
        _missing_coverage.append(f"PDF column-alignment test skipped: {SYMMETRY_INSPECTION.name} absent")
        print(f"  [MISSING] {SYMMETRY_INSPECTION.name} absent")
        return

    pages = ce.read_pdf_layouts(SYMMETRY_INSPECTION)
    check(bool(pages), "rendered at least one page", f"{len(pages)} page(s) with text")
    page1 = pages[0][1]

    header = next((l for l in page1.splitlines() if "XS" in l and "2XL" in l), None)
    ship = next((l for l in page1.splitlines() if "SHIPPMENT" in l and "BLACK" in l), None)
    check(header is not None, "found the size header line")
    check(ship is not None, "found the BLACK shipment line")
    if not (header and ship):
        return

    # The whole point: '19' must sit under 'XS', not under 'S'. Flattened text
    # loses this and a reader would have to guess.
    xs_col, s_col = header.index("XS"), header.index(" S ") + 1
    v19 = ship.index("19")
    check(abs(v19 - xs_col) <= 2, "BLACK's first figure aligns with the XS column", f"19@{v19} vs XS@{xs_col}")
    check(abs(v19 - s_col) > 2, "and is NOT aligned with the S column", f"19@{v19} vs S@{s_col}")

    twoxl_col = header.index("2XL")
    check(
        len(ship.rstrip()) < twoxl_col or not ship[twoxl_col : twoxl_col + 3].strip().isdigit(),
        "2XL column is blank on the shipment row",
    )
    check("110" in ship and "321" in ship, "M figure and colour total both present on the row")

    # All 5 pages of this report render: the photo/measurement pages are not
    # text-free, they carry a heading ("MEASUREMENT"). They cost almost nothing
    # to include, and the model reports them as holding no line data.
    sources, warnings = dp.build_source_documents([SYMMETRY_INSPECTION])
    check(len(sources) == 5, "all pages with any text layer are sent", f"{len(sources)} of 5")
    check(warnings == [], "nothing to warn about when every page has text", str(warnings))
    check(
        all(s.label.startswith(SYMMETRY_INSPECTION.name) for s in sources),
        "each source labelled with its file and page so source_hint is traceable",
        sources[0].label,
    )

    # A genuinely text-free page (a scan) must be REPORTED, not silently dropped
    # -- it could be the page holding the size table. Simulate by having the
    # renderer return fewer pages than the PDF has.
    saved = ce.read_pdf_layouts
    try:
        ce.read_pdf_layouts = lambda p: [("page 1", "some text")]
        _sources, warnings = dp.build_source_documents([SYMMETRY_INSPECTION])
        check(
            any("no text layer" in w and "4 of 5" in w for w in warnings),
            "pages dropped for having no text layer are reported with a count",
            str(warnings)[:100],
        )
        ce.read_pdf_layouts = lambda p: []
        _sources, warnings = dp.build_source_documents([SYMMETRY_INSPECTION])
        check(
            any("NOT included" in w for w in warnings),
            "a fully unreadable PDF says its content was excluded",
            str(warnings)[:100],
        )
    finally:
        ce.read_pdf_layouts = saved


def test_multidoc_call_shape(tmp: Path) -> None:
    section("multi-document extraction call shape")
    inprotex = tmp / "inprotex_like.xlsx"
    unknown = tmp / "unknown_vendor.xlsx"
    write_inprotex_like(inprotex)
    write_unknown_vendor(unknown)

    client, parse = fake_client([response(packing(vendor="Two Doc Vendor"))])
    result = dp.parse_shipment_documents(
        [inprotex, unknown], focus="PO 1721, style W600001", extractor=ce.ClaudeExtractor(client=client)
    )

    check(result.parser == "claude-assisted-multidoc", "reports the multidoc parser", result.parser)
    check(len(parse.calls) == 1, "ONE call with both documents (so they can be cross-referenced)")

    kwargs = parse.calls[0]
    system = kwargs["system"][0]
    check(system["text"] is ce.MULTI_DOC_SYSTEM_PROMPT, "uses the multi-document system prompt")
    check(system["cache_control"] == {"type": "ephemeral"}, "multi-doc prompt is cached too")
    check(
        "column header it sits under" in system["text"],
        "prompt forbids reading figures by left-to-right position",
    )
    check("SHIPPED quantities, not ordered" in system["text"], "prompt says shipped, not ordered")

    blocks = kwargs["messages"][0]["content"]
    texts = [b["text"] for b in blocks]
    check(any("Extract ONLY: PO 1721" in t for t in texts), "focus scope passed to the model")
    check(sum("===== DOCUMENT" in t for t in texts) == 2, "one delimited block per source document")
    check(any("2 document(s) for this one shipment" in t for t in texts), "manifest lists the sources")

    check(any("combined 2 rendered source" in n for n in result.notes), "sources recorded in notes")
    check(any("scope limited to" in n for n in result.notes), "focus recorded in notes")

    expect_raises(
        ce.ExtractionError,
        lambda: ce.ClaudeExtractor(client=client).extract_documents([]),
        "empty source list raises",
    )
    expect_raises(
        FileNotFoundError,
        lambda: dp.parse_shipment_documents([tmp / "nope.pdf"]),
        "missing source file raises",
    )
    bad = tmp / "notes.txt"
    bad.write_text("x", encoding="utf-8")
    expect_raises(
        ce.ExtractionError,
        lambda: dp.parse_shipment_documents([bad]),
        "unsupported file type raises",
    )


def test_attachment_classifier_offline(tmp: Path) -> None:
    section("attachment triage — filename rules (offline)")
    import attachment_classifier as ac

    def ftype(name: str):
        doc_type, ambiguous, _reason = ac.classify_by_filename(name)
        return doc_type, ambiguous

    check(ftype("FA26 7TH W600001 PO1721 FINAL INSPECTION REPORT.pdf")[0] == ac.DocType.INSPECTION_REPORT,
          "inspection report recognised by name")
    check(ftype("SD Vendor Payment Request #9001, 9002.pdf")[0] == ac.DocType.PAYMENT_REQUEST,
          "payment request recognised")
    check(ftype("SD Ocean Schedule 1720.pdf")[0] == ac.DocType.SHIPPING_SCHEDULE, "ocean schedule recognised")
    check(ftype("Shipping Advice 6128990769.pdf")[0] == ac.DocType.SHIPPING_ADVICE, "shipping advice recognised")
    check(ftype("SD Actual Packing Covering ^N1720^J 1721.pdf")[0] == ac.DocType.PACKING_LIST,
          "'actual packing' recognised as the packing list")

    # Both words present -> flagged ambiguous, because filenames genuinely cannot
    # settle it (the Symmetry invoice and the Inprotex packing slip look alike).
    invoice_type, invoice_ambiguous = ftype("SD #1720, 1721 INVOICE, PACKING LIST.pdf")
    check(invoice_ambiguous, "'INVOICE, PACKING LIST' flagged ambiguous, not decided by name")
    inpro_type, inpro_ambiguous = ftype("0626 (SD-219) Invoice_Packing.xlsx")
    check(inpro_ambiguous, "'Invoice_Packing' also flagged ambiguous")
    check(
        invoice_type == inpro_type,
        "the two opposite-answer files are indistinguishable by name -- which is why content decides",
    )

    check(ac.looks_like_rollup("SD Actual Packing Covering 1720.pdf"), "'covering' marks a rollup")
    check(not ac.looks_like_rollup("SD Actual Packing 1720.pdf"), "plain 'actual packing' is not a rollup")

    # A filename can never establish size-level-ness on its own.
    inprotex_like = tmp / "inprotex_like.xlsx"
    write_inprotex_like(inprotex_like)
    res = ac.classify_attachments([inprotex_like], use_content_check=False)
    check(any("content check disabled" in w for w in res.warnings),
          "disabling the content check is reported as unreliable")

    # An inspection report is excluded without spending a content call on it.
    if SYMMETRY_INSPECTION.exists():
        res = ac.classify_attachments([SYMMETRY_INSPECTION], use_content_check=False)
        check(res.needs_manual_entry, "an email with only an inspection report -> manual entry")
        check(res.excluded and "Paula" in res.excluded[0].excluded_reason,
              "excluded on the standing ruling")
        check(res.excluded[0].preview_chars == 0, "no preview built for it -- no content call spent")

    check(ac.DocType.INSPECTION_REPORT in ac.BANNED_AS_DATA_SOURCE, "ban is declared in code, not just prose")

    # The guard on the cross-document join.
    if SYMMETRY_INSPECTION.exists() and SYMMETRY_COVERING.exists():
        expect_raises(
            ce.ExtractionError,
            lambda: dp.parse_shipment_documents([SYMMETRY_COVERING, SYMMETRY_INSPECTION]),
            "parse_shipment_documents REFUSES an inspection report",
        )


def test_sheet_selection(tmp: Path) -> None:
    section("multi-sheet workbook: extract ONLY the packing sheet")
    multi = HERE / "fixtures" / "multi_sheet_one_packing.xlsx"
    none_ = HERE / "fixtures" / "no_packing_sheet.xlsx"
    if not (multi.exists() and none_.exists()):
        _missing_coverage.append(
            "sheet-selection tests skipped: run `python make_test_fixtures.py` to build fixtures"
        )
        print("  [MISSING] fixtures absent -- run make_test_fixtures.py")
        return

    import attachment_classifier as ac

    def verdicts(*specs):
        """A mocked _ContentVerdicts response for the classification call."""
        return response(
            ac._ContentVerdicts(
                verdicts=[
                    ac._ContentVerdict(doc_type=dt, has_size_breakdown=sz, reason=why)
                    for dt, sz, why in specs
                ]
            )
        )

    # --- the defect this fixes: 4 sheets, only 1 is a packing list --------------
    grids = [g for g in ce.read_workbook_grids(multi) if not g.is_empty]
    names = [g.name for g in grids]
    check(len(grids) == 4, "fixture has 4 non-empty sheets", str(names))
    check(names[-1] == "PACKING", "the packing sheet is LAST (a first-sheet fix would fail)", names[-1])

    client, parse = fake_client([
        verdicts(
            ("commercial_invoice", False, "COMMERCIAL INVOICE header, quantities at colour level"),
            ("other", False, "shipment/vessel details, no line data"),
            ("commercial_invoice", False, "divided freight charges, no sizes"),
            ("packing_list", True, "PACKING LIST with S/M/L size columns"),
        ),
        response(packing(lines=[line(po="9001", style="S000001", color="BLK", size="S", qty=100)])),
    ])
    result = dp.parse_packing_slip(multi, extractor=ce.ClaudeExtractor(client=client), force="claude")

    check(len(parse.calls) == 2, "exactly 2 calls: ONE classification + ONE extraction", str(len(parse.calls)))
    check(
        parse.calls[0]["system"][0]["text"] is ac.CLASSIFIER_SYSTEM_PROMPT,
        "classification reuses the EXISTING classifier prompt, not a new one",
    )
    check(
        parse.calls[0]["output_format"] is ac._ContentVerdicts,
        "and the existing classifier schema",
    )
    sections = [b["text"] for b in parse.calls[0]["messages"][0]["content"]]
    check(sum("===== SECTION" in t for t in sections) == 4, "all 4 sheets classified in that one call")

    extraction_text = parse.calls[1]["messages"][0]["content"][0]["text"]
    check("PACKING LIST" in extraction_text or "PACKING" in extraction_text,
          "the extraction call received the packing sheet")
    check("COMMERCIAL INVOICE" not in extraction_text,
          "and did NOT receive the invoice sheet -- the actual defect")
    check("DIVIDED CHARGES" not in extraction_text, "nor the divided-charges sheet")
    check("VESSEL" not in extraction_text, "nor the shipment sheet")

    # Requirement 4: every skipped sheet logged with type and reason.
    skips = [w for w in result.warnings if "NOT extracted" in w]
    check(len(skips) == 3, "one skip note per non-packing sheet", str(len(skips)))
    for expected in ("INVOICE(PAYMENT)", "SHIPMENT", "(+DIV CHARGE)."):
        check(any(expected in w for w in skips), f"skip note names sheet {expected}")
    check(all("predicted type:" in w for w in skips), "each skip note carries the predicted type")
    check(all("reason:" in w for w in skips), "each skip note carries a reason")

    # --- requirement 3: no packing sheet anywhere -> explicit, not silent ------
    client, parse = fake_client([
        verdicts(
            ("commercial_invoice", False, "COMMERCIAL INVOICE, no size columns"),
            ("other", False, "totals only, no line data"),
        )
    ])
    try:
        dp.parse_packing_slip(none_, extractor=ce.ClaudeExtractor(client=client), force="claude")
        check(False, "no-packing-sheet workbook raises", "returned instead of raising")
    except ce.NoPackingSheetFound as exc:
        check(True, "raises NoPackingSheetFound rather than returning zero lines")
        text = str(exc)
        check(none_.name in text, "message names the workbook", none_.name)
        check("'INVOICE'" in text and "'SUMMARY'" in text, "message lists every sheet examined")
        check("commercial_invoice" in text, "with each sheet's predicted type")
        check("NOT an empty shipment" in text, "and says explicitly it is not an empty shipment")
        check(len(exc.verdicts) == 2, "verdicts retained on the exception for programmatic use")
    except Exception as exc:  # noqa: BLE001
        check(False, "raises NoPackingSheetFound", f"raised {type(exc).__name__}: {exc}")
    check(len(parse.calls) == 1, "only the classification call was spent -- no extraction attempted")
    check(issubclass(ce.NoPackingSheetFound, ce.ExtractionError),
          "NoPackingSheetFound is an ExtractionError, so existing handlers still catch it")

    # --- single-sheet workbooks cost nothing extra -----------------------------
    single = tmp / "inprotex_like.xlsx"
    write_inprotex_like(single)
    client, parse = fake_client([response(packing())])
    dp.parse_packing_slip(single, extractor=ce.ClaudeExtractor(client=client), force="claude")
    check(len(parse.calls) == 1, "single-sheet workbook makes NO classification call", str(len(parse.calls)))

    # --- an explicit sheet_names filter is the caller's decision, not re-judged -
    client, parse = fake_client([response(packing())])
    ce.ClaudeExtractor(client=client).extract_workbook(multi, sheet_names=["PACKING"])
    check(len(parse.calls) == 1, "explicit sheet_names skips classification", str(len(parse.calls)))

    # --- the selector's own contract -------------------------------------------
    sc_keep = ac.SectionClassification("PACKING", ac.DocType.PACKING_LIST, True, "has sizes")
    sc_drop = ac.SectionClassification("INVOICE", ac.DocType.COMMERCIAL_INVOICE, False, "no sizes")
    sc_sizeless = ac.SectionClassification("PL", ac.DocType.PACKING_LIST, False, "no size cols")
    sc_banned = ac.SectionClassification("QC", ac.DocType.INSPECTION_REPORT, True, "has sizes")
    check(sc_keep.is_shipment_data, "a packing-list section with sizes is extracted")
    check(not sc_drop.is_shipment_data, "an invoice section is not")
    check(not sc_sizeless.is_shipment_data, "a packing list without sizes is not")
    check(not sc_banned.is_shipment_data,
          "an inspection-report section is not, even WITH sizes (Paula's ruling holds at sheet level)")
    check("never a shipment-data source" in sc_banned.skip_reason, "and says why")
    check("not a packing list" in sc_drop.skip_reason, "invoice skip reason is specific")
    check("no per-size quantities" in sc_sizeless.skip_reason, "sizeless skip reason is specific")

    # sheet_preview is shared with the whole-file preview, not a second implementation
    pk = next(g for g in grids if g.name == "PACKING")
    prev = ac.sheet_preview(pk)
    check("PACKING" in prev and "S" in prev, "sheet_preview renders the sheet")
    check(ac._find_size_header_row(pk) is not None, "reuses the existing size-header seek")


def test_size_header_canonical(tmp: Path) -> None:
    section("size-header detection handles full-width and spaced labels")
    import openpyxl

    import attachment_classifier as ac
    from canonical import canonical

    def sheet_with(header_cells, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PACKING"
        ws.append(["SAMPLE VENDOR CO., LTD."])
        ws.append(["PACKING LIST"])
        for _ in range(18):  # push the header below the preview window
            ws.append([None])
        ws.append(["C/NO.", "COLOR"] + list(header_cells) + ["TOTAL"])
        ws.append([1, "BLK", 10, 20, 30, 60])
        wb.save(path)
        return [g for g in ce.read_workbook_grids(path) if not g.is_empty][0]

    # ASCII baseline -- must still work.
    g = sheet_with(["S", "M", "L"], tmp / "hdr_ascii.xlsx")
    check(ac._find_size_header_row(g) is not None, "ASCII size header still found",
          str(ac._find_size_header_row(g)))

    # THE POINT: full-width size labels. Before canonicalization these were
    # invisible to the detector, so a real packing sheet could look sizeless --
    # which now cascades to NoPackingSheetFound and manual entry.
    g = sheet_with(["\uff33", "\uff2d", "\uff2c"], tmp / "hdr_fullwidth.xlsx")
    check(ac._find_size_header_row(g) is not None,
          "FULL-WIDTH size header (\uff33\uff2d\uff2c) is found", str(ac._find_size_header_row(g)))

    g = sheet_with(["\uff12\uff38", "\uff13\uff38", "XL"], tmp / "hdr_fw_2x.xlsx")
    check(ac._find_size_header_row(g) is not None,
          "full-width 2X/3X recognised", str(ac._find_size_header_row(g)))

    # Spacing and case variants.
    g = sheet_with(["  S  ", "m", "One  Size"], tmp / "hdr_spaced.xlsx")
    check(ac._find_size_header_row(g) is not None,
          "padded / lowercase / double-spaced 'One  Size' recognised",
          str(ac._find_size_header_row(g)))

    # Negative control: a row of unrelated words must NOT look like a size header.
    g = sheet_with(["WIDTH", "HEIGHT", "DEPTH"], tmp / "hdr_none.xlsx")
    check(ac._find_size_header_row(g) is None, "unrelated headers are not mistaken for sizes",
          str(ac._find_size_header_row(g)))

    check(canonical("\uff12\uff38") in ac._SIZE_TOKENS_CANON,
          "the canonical token set contains the folded form of 2X")
    check(len(ac._SIZE_TOKENS_CANON) == len({canonical(t) for t in ac._SIZE_TOKENS}),
          "canonical token set is derived from _SIZE_TOKENS, not duplicated by hand")



def test_canonical_form(tmp: Path) -> None:
    section("canonical form covers the CLASS, not just the double space")
    from canonical import canonical, canonical_key, same

    # Each pair must canonicalise identically. These are the members of the class
    # named in the design: whitespace, full-width forms, dashes, case, CJK.
    equivalent = [
        ("double internal space", "NEW  INDIGO", "NEW INDIGO"),
        ("triple internal space", "NEW   INDIGO", "NEW INDIGO"),
        ("non-breaking space", "NEW INDIGO", "NEW INDIGO"),
        ("narrow no-break space", "NEW INDIGO", "NEW INDIGO"),
        ("ideographic space", "TID　BLK", "TID BLK"),
        ("zero-width space", "NEW​INDIGO", "NEW INDIGO"),
        ("full-width comma", "PO#1657，M630018", "PO#1657,M630018"),
        ("full-width digits+letter", "２Ｘ", "2X"),
        ("full-width alpha", "ＢＬＫ", "BLK"),
        ("en-dash vs hyphen", "32–34", "32-34"),
        ("em-dash vs hyphen", "32—34", "32-34"),
        ("minus sign vs hyphen", "32−34", "32-34"),
        ("non-breaking hyphen", "32‑34", "32-34"),
        ("mixed case", "new indigo", "NEW INDIGO"),
        ("leading/trailing space", "   NEW INDIGO   ", "NEW INDIGO"),
        ("tab and newline", "NEW\tINDIGO\n", "NEW INDIGO"),
        ("CJK with double space", "建跍  空運", "建跍 空運"),
        ("CJK padded", "  建跍空運  ", "建跍空運"),
    ]
    for label, a, b in equivalent:
        check(same(a, b), f"equivalent: {label}", f"{canonical(a)!r}")

    # Genuinely different values must NOT be conflated -- a normalizer that maps
    # everything together would "fix" the bug by breaking matching entirely.
    distinct = [
        ("different colour", "BLACK", "BLUE"),
        ("different size", "32-34", "32-36"),
        ("different length", "M", "MM"),
        ("different style", "M120246", "M120247"),
        ("substring", "TID", "TIDE"),
        ("size vs waist-inseam", "32", "32-34"),
    ]
    for label, a, b in distinct:
        check(not same(a, b), f"distinct: {label}", f"{canonical(a)!r} vs {canonical(b)!r}")

    check(canonical(None) == "", "None canonicalises to empty string")
    check(canonical("") == "", "empty string stays empty")
    check(canonical("   ") == "", "whitespace-only becomes empty")
    check(canonical(1662) == "1662", "non-str input is coerced")
    check(canonical_key("A", "b ") == ("a", "b"), "canonical_key canonicalises each part",
          str(canonical_key("A", "b ")))


def test_verbatim_source_preserved(tmp: Path) -> None:
    section("verbatim source text survives canonicalization untouched")
    from canonical import canonical
    from extraction_schema import aggregate_lines, line_to_dict, ExtractedLine

    dirty = "NEW  INDIGO"  # double space, as the real Symmetry PDF prints it

    # line_to_dict must not rewrite the value it was given.
    d = line_to_dict(ExtractedLine(
        po_number="1720", style_number="M650022", color=dirty, size="S",
        quantity=22, confidence="high", note="", source_hint="P1!R7",
    ))
    check(d["color"] == dirty, "line_to_dict keeps the colour byte-for-byte", repr(d["color"]))
    check(canonical(d["color"]) == "new indigo", "while its canonical form is clean")

    # Aggregation groups on the canonical key but must not overwrite display text.
    rows = [
        {"po_number": "1720", "style_number": "M650022", "color": dirty, "size": "S",
         "quantity": 10, "confidence": "high", "note": "", "source_hint": "R7"},
        {"po_number": "1720", "style_number": "M650022", "color": "NEW INDIGO", "size": "S",
         "quantity": 12, "confidence": "high", "note": "", "source_hint": "R8"},
    ]
    out, warns = aggregate_lines(rows)
    check(len(out) == 1, "the two whitespace variants collapsed to one line", str(len(out)))
    check(out[0]["quantity"] == 22, "quantities summed across the variants", str(out[0]["quantity"]))
    check(out[0]["color"] in (dirty, "NEW INDIGO"),
          "the surviving colour is one of the VERBATIM variants, not the canonical form",
          repr(out[0]["color"]))
    check(out[0]["color"] != canonical(out[0]["color"]) or out[0]["color"] == "NEW INDIGO",
          "it is not lowercased")
    check("printed as" in out[0]["note"], "the note records that the source used both renderings",
          out[0]["note"][:90])
    check(repr(dirty) in out[0]["note"] and repr("NEW INDIGO") in out[0]["note"],
          "and quotes both verbatim variants")

    # Determinism: whichever order the rows arrive in, the surviving display value
    # and the row order are the same.
    out2, _ = aggregate_lines(list(reversed(rows)))
    check(out2[0]["color"] == out[0]["color"],
          "the surviving variant is chosen deterministically, not by input order")


def test_matcher_canonical_both_sides(tmp: Path) -> None:
    section("matcher canonicalises BOTH operands (NetSuite side may be dirty too)")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    def ns(color="NEW INDIGO", size="M", qty=100, style="M650022"):
        return POLine(
            line_id="10", item=f"{style} : {style}-{color}-{size}", style_number=style,
            vendor_name=None, color=color, size=size, quantity=qty, units="Ea",
            expected_receipt_date=dt.date(2026, 9, 1), override_expected_receipt=False,
            updated_receipt_date=None,
        )

    def vendor(color="NEW INDIGO", size="M", qty=110, style="M650022"):
        return {"po_number": "1720", "style_number": style, "color": color, "size": size,
                "quantity": qty, "confidence": "high", "note": ""}

    # 1. The original bug: vendor side dirty, NetSuite clean.
    client = NetSuiteClient(mock_data={"1720": [ns(color="NEW INDIGO")]})
    c = mt.build_proposed_changes([vendor(color="NEW  INDIGO")], client)[0]
    check(c.status == mt.STATUS_PENDING_REVIEW,
          "vendor 'NEW  INDIGO' matches NetSuite 'NEW INDIGO'", c.status)
    check(c.current_quantity == 100 and c.proposed_quantity == 110, "and the diff is computed")

    # 2. The reverse: NetSuite side dirty, vendor clean. Normalizing only the
    #    extracted side would just relocate the mismatch.
    client = NetSuiteClient(mock_data={"1720": [ns(color="NEW  INDIGO")]})
    c = mt.build_proposed_changes([vendor(color="NEW INDIGO")], client)[0]
    check(c.status == mt.STATUS_PENDING_REVIEW,
          "DIRTY NETSUITE side still matches a clean vendor value", c.status)

    # 3. Both sides dirty, differently.
    client = NetSuiteClient(mock_data={"1720": [ns(color="new indigo")]})
    c = mt.build_proposed_changes([vendor(color="NEW  INDIGO")], client)[0]
    check(c.status == mt.STATUS_PENDING_REVIEW,
          "nbsp on one side and double space on the other still match", c.status)

    # 4. Sizes: dash variants and aliases, on either side.
    client = NetSuiteClient(mock_data={"1720": [ns(size="32-34")]})
    c = mt.build_proposed_changes([vendor(size="32–34")], client)[0]
    check(c.status == mt.STATUS_PENDING_REVIEW, "en-dash size matches hyphen size", c.status)
    client = NetSuiteClient(mock_data={"1720": [ns(size="2X")]})
    for label, vend in (("XXL", "XXL"), ("2XL", "2XL"), ("lowercase 2x", "2x"),
                        ("padded", "  2XL  ")):
        c = mt.build_proposed_changes([vendor(size=vend)], client)[0]
        check(c.status == mt.STATUS_PENDING_REVIEW, f"vendor size {label} matches NetSuite 2X", c.status)

    # 5. Style with a full-width form.
    client = NetSuiteClient(mock_data={"1720": [ns(style="M630018")]})
    c = mt.build_proposed_changes([vendor(style="Ｍ630018")], client)[0]
    check(c.status == mt.STATUS_PENDING_REVIEW, "full-width style letter matches ASCII", c.status)

    # 6. Genuinely different values must still NOT match -- canonicalization must
    #    not turn a miss into a false positive.
    client = NetSuiteClient(mock_data={"1720": [ns(color="BLACK")]})
    c = mt.build_proposed_changes([vendor(color="BLUE")], client)[0]
    check(c.status == mt.STATUS_NEEDS_ATTENTION, "BLACK vs BLUE still does NOT match", c.status)
    client = NetSuiteClient(mock_data={"1720": [ns(size="32-34")]})
    c = mt.build_proposed_changes([vendor(size="32-36")], client)[0]
    check(c.status == mt.STATUS_NEEDS_ATTENTION, "32-34 vs 32-36 still does NOT match", c.status)

    # 7. _normalize_size keeps returning NetSuite's own casing for display, while
    #    _size_key is the comparison form.
    check(mt._normalize_size("XXL") == "2X", "display label unchanged: XXL -> 2X", mt._normalize_size("XXL"))
    check(mt._normalize_size("  xxl  ") == "2X", "and is robust to case/padding now")
    check(mt._size_key("XXL") == mt._size_key("2X") == "2x",
          "comparison key folds both to '2x'", mt._size_key("XXL"))

    # 8. unmatched_netsuite_lines uses the same key on both sides.
    lines = [ns(color="NEW  INDIGO", size="M"), ns(color="BLACK", size="M")]
    left = mt.unmatched_netsuite_lines([vendor(color="NEW INDIGO", size="M")], lines)
    check([l.color for l in left] == ["BLACK"],
          "a dirty NetSuite colour is recognised as shipped, not reported unmatched",
          str([l.color for l in left]))


def test_line_aggregation(tmp: Path) -> None:
    section("aggregation by semantic key (determinism on retry)")
    from extraction_schema import aggregate_lines

    def ln(po="1721", style="W600001", color="BLACK", size="M", qty=10,
           conf="high", note="", hint="P1!R1"):
        return {"po_number": po, "style_number": style, "color": color, "size": size,
                "quantity": qty, "confidence": conf, "note": note, "source_hint": hint}

    # --- duplicate keys collapse, quantities sum -------------------------------
    out, warns = aggregate_lines([ln(qty=10, hint="P1!R5"), ln(qty=12, hint="P1!R9"), ln(size="L", qty=7)])
    check(len(out) == 2, "two duplicate rows collapsed into one, distinct key kept", str(len(out)))
    merged = next(l for l in out if l["size"] == "M")
    check(merged["quantity"] == 22, "quantities summed (10 + 12)", str(merged["quantity"]))
    check("P1!R5" in merged["source_hint"] and "P1!R9" in merged["source_hint"],
          "both source hints preserved", merged["source_hint"])
    check(any("aggregated" in w for w in warns), "collapse reported in warnings")

    # --- confidence takes the MINIMUM; a collapse must never upgrade -----------
    out, _ = aggregate_lines([ln(qty=5, conf="high"), ln(qty=5, conf="low")])
    check(out[0]["confidence"] == "low", "high + low -> low (worst wins)", out[0]["confidence"])
    out, _ = aggregate_lines([ln(qty=5, conf="high"), ln(qty=5, conf="medium")])
    check(out[0]["confidence"] == "medium", "high + medium -> medium", out[0]["confidence"])
    out, _ = aggregate_lines([ln(qty=5, conf="high"), ln(qty=5, conf="high")])
    check(out[0]["confidence"] == "high", "high + high stays high", out[0]["confidence"])
    out, _ = aggregate_lines([ln(qty=5, conf="low"), ln(qty=5, conf="high")])
    check(out[0]["confidence"] == "low", "order of inputs does not change the verdict", out[0]["confidence"])

    # notes unioned, deduplicated
    out, _ = aggregate_lines([ln(qty=1, note="carton 1 unlabelled"), ln(qty=2, note="carton 1 unlabelled"),
                              ln(qty=3, note="colour carried forward")])
    check(out[0]["note"].count("carton 1 unlabelled") == 1, "duplicate notes deduplicated", out[0]["note"])
    check("colour carried forward" in out[0]["note"], "distinct notes both kept", out[0]["note"])

    # --- sizeless rows are NOT collapsed --------------------------------------
    sizeless = [ln(size="", qty=1217, color="BLK", hint="INV!R23"),
                ln(size="", qty=1216, color="IND", hint="INV!R24"),
                ln(size="", qty=305, color="CBR", hint="INV!R26")]
    out, warns = aggregate_lines(sizeless)
    check(len(out) == 3, "three sizeless rows stay three rows", str(len(out)))
    check(sum(l["quantity"] for l in out) == 1217 + 1216 + 305, "their quantities are untouched")
    check(any("no size" in w for w in warns), "reported as not-aggregated", str(warns)[:80])

    # Two sizeless rows with an OTHERWISE identical key must still not merge --
    # a blank size cannot be a merge key.
    same = [ln(size="", qty=5), ln(size="", qty=6)]
    out, _ = aggregate_lines(same)
    check(len(out) == 2, "identical-but-sizeless rows are not merged into one bucket", str(len(out)))
    check({l["quantity"] for l in out} == {5, 6}, "and keep their separate quantities")

    # --- deterministic ordering ------------------------------------------------
    import random

    base = [ln(po="1721", style="W600001", color="BLACK", size="XS", qty=19),
            ln(po="1720", style="M650022", color="NEW INDIGO", size="S", qty=22),
            ln(po="1721", style="W520005", color="COCONUT", size="L", qty=33),
            ln(po="1721", style="W600001", color="BLACK", size="M", qty=110),
            ln(po="1720", style="M650022", color="NEW INDIGO", size="3XL", qty=4)]
    reference = [(l["po_number"], l["style_number"], l["color"], l["size"])
                 for l in aggregate_lines(list(base))[0]]
    stable = True
    for seed in range(6):
        shuffled = list(base)
        random.Random(seed).shuffle(shuffled)
        got = [(l["po_number"], l["style_number"], l["color"], l["size"])
               for l in aggregate_lines(shuffled)[0]]
        if got != reference:
            stable = False
    check(stable, "same input in ANY order yields the same output order", str(reference[:2]))
    check(reference == sorted(reference), "output is sorted by the semantic key")

    # --- edge cases ------------------------------------------------------------
    out, warns = aggregate_lines([])
    check(out == [] and warns == [], "empty input is a clean no-op")
    out, warns = aggregate_lines([ln()])
    check(len(out) == 1 and warns == [], "a single row produces no collapse warning")
    # SUPERSEDED BY CHANGE 4 (canonicalization). This previously asserted that
    # "BLACK" and "black" stay apart, on the reasoning that keeping them separate
    # was safer. Canonicalization casefolds, so they now collapse -- which is the
    # correct behaviour: the matcher already compared colour case-insensitively,
    # so treating them as different rows only ever produced two proposed changes
    # for one NetSuite line. The verbatim renderings are still both recorded.
    out, _ = aggregate_lines([ln(color="BLACK", qty=1), ln(color="black", qty=2)])
    check(len(out) == 1, "case variants of one colour now collapse (canonical key)", str(len(out)))
    check(out[0]["quantity"] == 3, "with quantities summed", str(out[0]["quantity"]))
    check(out[0]["color"] in ("BLACK", "black"), "displaying a verbatim variant, not the canonical form",
          repr(out[0]["color"]))
    check("printed as" in out[0]["note"], "and recording that both renderings appeared")
    # Genuinely different colours must still stay apart.
    out, _ = aggregate_lines([ln(color="BLACK", qty=1), ln(color="BLUE", qty=2)])
    check(len(out) == 2, "different colours are still not merged", str(len(out)))


def test_aggregation_wired_into_parsers(tmp: Path) -> None:
    section("aggregation is wired into BOTH parser paths")
    inprotex = tmp / "inprotex_like.xlsx"
    write_inprotex_like(inprotex)

    # The Claude path: a mocked response with a duplicate key must collapse.
    dupes = packing(lines=[
        line(po="1662", style="M120246", color="TID", size="S", qty=4, hint="R10"),
        line(po="1662", style="M120246", color="TID", size="S", qty=5, hint="R11"),
        line(po="1662", style="M120246", color="TID", size="M", qty=7, hint="R12"),
    ])
    client, _ = fake_client([response(dupes)])
    result = dp.parse_packing_slip(inprotex, extractor=ce.ClaudeExtractor(client=client), force="claude")
    check(len(result.lines) == 2, "Claude path collapsed the duplicate key", str(len(result.lines)))
    s_line = next(l for l in result.lines if l["size"] == "S")
    check(s_line["quantity"] == 9, "summed to 9 (4 + 5)", str(s_line["quantity"]))
    check(any("aggregated" in w for w in result.warnings), "collapse surfaced in ParseResult warnings")
    keys = [(l["po_number"], l["style_number"], l["color"], l["size"]) for l in result.lines]
    check(keys == sorted(keys), "Claude path output is sorted")

    # The deterministic path also aggregates and sorts, so both paths are
    # comparable and retry-stable.
    det = dp.parse_packing_slip(inprotex)
    check(det.parser == "inprotex-deterministic", "deterministic routing intact", det.parser)
    det_keys = [(l["po_number"], l["style_number"], l["color"], l["size"]) for l in det.lines]
    check(det_keys == sorted(det_keys), "deterministic path output is sorted too")
    check(len(det.lines) == 3, "3 lines from the fixture recap, unchanged", str(len(det.lines)))


def test_matcher_paula_rulings(tmp: Path) -> None:
    section("diff engine — Paula's rulings (2026-08-11)")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    def ns_line(size="M", qty=100, line_id="10"):
        return POLine(
            line_id=line_id, item=f"W600001 : W600001-Harper-BLACK-{size}",
            style_number="W600001", vendor_name=None, color="BLACK", size=size,
            quantity=qty, units="Ea", expected_receipt_date=dt.date(2026, 9, 1),
            override_expected_receipt=False, updated_receipt_date=None,
        )

    client = NetSuiteClient(mock_data={"1721": [ns_line("M", 100, "10"), ns_line("L", 50, "11")]})
    vendor = [{"po_number": "1721", "style_number": "W600001", "color": "BLACK", "size": "M",
               "quantity": 110, "confidence": "high", "note": ""}]

    changes = mt.build_proposed_changes(vendor, client, eta="2026/8/16", etd="2026/8/5")
    check(len(changes) == 1, "one record per vendor line")
    c = changes[0]

    # RULING 5 -- no proposed receipt dates.
    check(not hasattr(c, "proposed_expected_receipt_date"),
          "ProposedChange has NO proposed_expected_receipt_date field at all")
    check(not hasattr(c, "proposed_updated_receipt_date"), "and no proposed_updated_receipt_date")
    check(c.vendor_eta == "2026-08-16" and c.vendor_etd == "2026-08-05",
          "vendor ETD/ETA carried as reference", f"{c.vendor_etd} / {c.vendor_eta}")
    check(c.receipt_date_pending, "receipt date starts pending -- nobody has typed one")
    check("Reference only, not a proposed value" in c.reference_dates_label,
          "review UI label says reference, not proposal", c.reference_dates_label[:60])

    qty_only = c.to_netsuite_fields(include_dates=False)
    check(set(qty_only) == {"quantity"}, "quantity-only write contains ONLY quantity", str(set(qty_only)))
    check(qty_only["quantity"] == 110, "quantity replaces the ordered amount", str(qty_only["quantity"]))
    expect_raises(
        mt.DateNotConfirmed,
        lambda: c.to_netsuite_fields(include_dates=True),
        "writing dates before a human confirms one RAISES",
    )

    c.confirm_receipt_date("2026-09-10")
    with_dates = c.to_netsuite_fields(include_dates=True)
    check(with_dates["expectedReceiptDate"] == "2026-09-10", "confirmed date written to expectedReceiptDate")
    check(with_dates["custcol_sd_updatedreceiptdate"] == "2026-09-10", "and to the updated-receipt custom field")
    check(with_dates["custcol_override_expected_receipt"] is True, "override flag set alongside")
    check(
        with_dates["expectedReceiptDate"] != c.vendor_eta,
        "the written date is Paula's, NOT the vendor ETA",
        f"wrote {with_dates['expectedReceiptDate']}, vendor ETA was {c.vendor_eta}",
    )
    expect_raises(ValueError, lambda: c.confirm_receipt_date("09/10/2026"), "non-ISO confirmed date rejected")

    # RULING 6 -- over-shipment is normal and must not be flagged.
    over = mt.build_proposed_changes(
        [{**vendor[0], "quantity": 500}], client, eta="2026/8/16"
    )[0]
    check(over.status == mt.STATUS_PENDING_REVIEW,
          "shipped 500 vs ordered 100 is a plain PENDING_REVIEW", over.status)
    check(over.attention_reason == "", "no attention flag raised for over-shipment", repr(over.attention_reason))
    check(over.to_netsuite_fields(include_dates=False)["quantity"] == 500, "over-ship quantity replaces as-is")

    # RULING 7 -- a PO line absent from the packing list is a silent no-op.
    check(len(changes) == 1, "the untouched 'L' line produced NO change record")
    check(all(ch.size != "L" for ch in changes), "and no NO_CHANGE row either")
    leftover = mt.unmatched_netsuite_lines(vendor, client.get_purchase_order("1721"))
    check([l.size for l in leftover] == ["L"], "reportable for visibility only", str([l.size for l in leftover]))

    # Still flagged: a vendor line with no NetSuite counterpart is genuinely odd.
    orphan = mt.build_proposed_changes(
        [{**vendor[0], "size": "4XL"}], client, eta="2026/8/16"
    )[0]
    check(orphan.status == mt.STATUS_NEEDS_ATTENTION, "vendor line with no NetSuite match -> NEEDS_ATTENTION")
    check("no NetSuite line" in orphan.attention_reason, "reason names the mismatch")

    # Low-confidence extraction also earns attention.
    shaky = mt.build_proposed_changes(
        [{**vendor[0], "confidence": "low", "note": "quantity cell unreadable"}], client
    )[0]
    check(shaky.status == mt.STATUS_NEEDS_ATTENTION, "low-confidence line -> NEEDS_ATTENTION")
    check("confidence low" in shaky.attention_reason, "reason cites the confidence")

    # Whole shipment unresolvable -> everything needs attention (manual entry).
    manual = mt.build_proposed_changes(vendor, client, shipment_needs_manual_entry=True)[0]
    check(manual.status == mt.STATUS_NEEDS_ATTENTION, "manual-entry shipment marks lines NEEDS_ATTENTION")
    check("manual entry required" in manual.attention_reason, "reason says manual entry")

    # No change at all when the quantity already matches.
    same = mt.build_proposed_changes([{**vendor[0], "quantity": 100}], client)[0]
    check(same.status == mt.STATUS_NO_CHANGE, "matching quantity -> NO_CHANGE", same.status)
    check(same.to_netsuite_fields(include_dates=False) == {}, "and produces an empty write")

    # Size normalization still spans both conventions.
    check(mt._normalize_size("XXL") == "2X" and mt._normalize_size("2XL") == "2X",
          "XXL and 2XL both normalize to NetSuite's 2X")
    check(mt._normalize_size("3XL") == "3X" and mt._normalize_size("XXXL") == "3X",
          "3XL and XXXL both normalize to 3X")


def test_closed_po_line(tmp: Path) -> None:
    section("closed NetSuite PO lines are flagged, never written to")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    def line(size, qty, line_id, closed):
        return POLine(
            line_id=line_id, item=f"W600001 : W600001-Harper-BLACK-{size}",
            style_number="W600001", vendor_name=None, color="BLACK", size=size,
            quantity=qty, units="Ea", expected_receipt_date=dt.date(2026, 9, 1),
            override_expected_receipt=False, updated_receipt_date=None, closed=closed,
        )

    client = NetSuiteClient(
        mock_data={"1721": [line("M", 100, "10", True), line("L", 50, "11", False)]}
    )
    vendor = [
        {"po_number": "1721", "style_number": "W600001", "color": "BLACK", "size": "M",
         "quantity": 110, "confidence": "high", "note": ""},
        {"po_number": "1721", "style_number": "W600001", "color": "BLACK", "size": "L",
         "quantity": 60, "confidence": "high", "note": ""},
    ]
    changes = {c.size: c for c in mt.build_proposed_changes(vendor, client)}

    closed = changes["M"]
    check(closed.status == mt.STATUS_NEEDS_ATTENTION,
          "a vendor line matching a CLOSED NetSuite line flags instead of proposing", closed.status)
    check(closed.line_closed, "the change records that the line is closed")
    check("closed in NetSuite" in closed.attention_reason,
          "reason states the line is closed", closed.attention_reason[:80])
    check(closed.status != mt.STATUS_PENDING_REVIEW, "and is NOT a normal PENDING_REVIEW quantity change")

    # The write path must refuse it outright, even if something upstream tried.
    expect_raises(
        mt.LineClosed,
        lambda: closed.to_netsuite_fields(include_dates=False),
        "building a write for a closed line RAISES LineClosed",
    )
    closed.confirm_receipt_date("2026-09-10")
    expect_raises(
        mt.LineClosed,
        lambda: closed.to_netsuite_fields(include_dates=True),
        "still refuses even with a confirmed date",
    )

    # An open line alongside it is unaffected -- one closed line must not poison
    # the rest of the shipment.
    open_line = changes["L"]
    check(open_line.status == mt.STATUS_PENDING_REVIEW, "the open line still proposes normally", open_line.status)
    check(not open_line.line_closed, "and is not marked closed")
    check(open_line.to_netsuite_fields(include_dates=False) == {"quantity": 60},
          "and builds its write fine", str(open_line.to_netsuite_fields(include_dates=False)))


def test_duplicate_key_resolution(tmp: Path) -> None:
    section("one key -> several NetSuite lines (never pick silently, never sum)")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    # Shapes taken from real sandbox POs. The values are reproduced; the records
    # themselves are synthetic so the suite stays offline.
    def ns(line_id, qty, *, is_open=True, recv=0.0, billed=0.0, closed=False,
           erd=dt.date(2026, 9, 1), override=False, upd=None, rate=12.5,
           style="A320001", color="WHT", size="ALL"):
        return POLine(
            line_id=line_id, item=f"{style} : {style}-{color}-{size}",
            style_number=style, vendor_name=None, color=color, size=size,
            quantity=qty, units="Ea", expected_receipt_date=erd,
            override_expected_receipt=override, updated_receipt_date=upd,
            closed=closed, is_open=is_open, quantity_received=recv,
            quantity_billed=billed, rate=rate,
        )

    def vendor(qty, style="A320001", color="WHT", size="ALL", po="1649"):
        return {"po_number": po, "style_number": style, "color": color, "size": size,
                "quantity": qty, "confidence": "high", "note": ""}

    def one(lines, vl):
        return mt.build_proposed_changes([vl], NetSuiteClient(mock_data={"1649": lines}))[0]

    # -- PO0001620 shape: two lines, same key, different dates, override differs.
    #    One is open. That one is the target, and the date fields play no part in
    #    choosing it.
    c = one(
        [ns("6", 6, is_open=False, erd=dt.date(2025, 10, 6), override=True,
            upd=dt.date(2025, 10, 6)),
         ns("42", 5, is_open=True, erd=dt.date(2025, 9, 16), override=False)],
        vendor(11),
    )
    check(c.status == mt.STATUS_PENDING_REVIEW,
          "PO0001620 shape: the single OPEN line of the pair is targeted", c.status)
    check(c.line_id == "42", "and it is the open one, not simply the first", str(c.line_id))
    check(c.current_quantity == 5,
          "current quantity is that ONE line's, never the pair summed", str(c.current_quantity))
    check(c.candidate_lines == [], "a resolved match carries no candidate payload")

    # -- PO0001514 shape: same date on both lines, quantities 20 and 3. Only the
    #    open one can be written to; the closed twin is not added to it.
    c = one(
        [ns("3", 20, is_open=False, recv=20.0, billed=20.0),
         ns("31", 3, is_open=True)],
        vendor(9),
    )
    check(c.status == mt.STATUS_PENDING_REVIEW,
          "PO0001514 shape: identical dates, one open line -> normal change", c.status)
    check(c.line_id == "31" and c.current_quantity == 3,
          "targets the open line; 20 + 3 is never staged as 23", f"{c.line_id}/{c.current_quantity}")
    check(c.to_netsuite_fields(include_dates=False) == {"quantity": 9},
          "and the write is a plain quantity update on that one line",
          str(c.to_netsuite_fields(include_dates=False)))

    # -- PO0001649: the genuinely ambiguous one. 50 units received 0, 200 units
    #    received 100, both open, both due the same day. The tool must not choose.
    ambiguous = [
        ns("1", 50, is_open=True, recv=0.0, billed=0.0, erd=dt.date(2026, 7, 1), override=False),
        ns("2", 200, is_open=True, recv=100.0, billed=100.0, erd=dt.date(2026, 7, 1), override=True),
    ]
    c = one(ambiguous, vendor(58))
    check(c.status == mt.STATUS_NEEDS_RESOLUTION,
          "PO0001649: two open lines -> NEEDS_RESOLUTION", c.status)
    check(c.line_id is None, "no line was picked", str(c.line_id))
    check(c.current_quantity is None,
          "and no line's quantity was adopted as 'current'", str(c.current_quantity))
    check(c.proposed_quantity == 58, "the vendor quantity is preserved verbatim for the reviewer")
    check("does not choose" in c.attention_reason and "1, 2" in c.attention_reason,
          "the reason names both line ids and says the tool will not choose",
          c.attention_reason[:90])

    # NEVER summed on the NetSuite side. This is the whole point of change 5:
    # extraction-side duplicates are the same style/colour/size counted twice in
    # one document (sum them), NetSuite-side duplicates are two separate PO lines
    # (never sum them).
    check(250 not in {c.current_quantity, c.proposed_quantity},
          "50 + 200 is nowhere in the staged change")
    check([p["quantity"] for p in c.candidate_lines] == [50, 200],
          "both candidate quantities are surfaced separately, not combined",
          str([p["quantity"] for p in c.candidate_lines]))

    # The payload is what a human decides from, so it has to carry the receipt
    # figures that distinguish the lines.
    payload = c.candidate_lines[1]
    for field_name, expected in (("line_id", "2"), ("quantity", 200), ("quantity_received", 100.0),
                                 ("quantity_billed", 100.0), ("expected_receipt_date", "2026-07-01"),
                                 ("override_expected_receipt", True), ("rate", 12.5),
                                 ("is_open", True)):
        check(payload.get(field_name) == expected,
              f"candidate payload carries {field_name}", f"{payload.get(field_name)!r}")
    check("updated_receipt_date" in payload, "candidate payload carries updated_receipt_date")

    # Scope boundary, asserted rather than trusted: custcol_sd_fg_excluderepspark
    # is Paula's field. This tool does not read, write or display it. (It also
    # failed as a discriminator -- it differs in only 25.5% of duplicate pairs.)
    check(all("repspark" not in k.lower() for p in c.candidate_lines for k in p),
          "the candidate payload does NOT carry custcol_sd_fg_excluderepspark")
    check("repspark" not in str(mt.asdict(c)).lower(),
          "nor does any other part of the staged change")

    # A write must be structurally impossible, not merely discouraged.
    expect_raises(
        mt.LineAmbiguous,
        lambda: c.to_netsuite_fields(include_dates=False),
        "building a write for an unresolved change RAISES LineAmbiguous",
    )
    c.confirm_receipt_date("2026-09-10")
    expect_raises(
        mt.LineAmbiguous,
        lambda: c.to_netsuite_fields(include_dates=True),
        "still refuses once a receipt date has been confirmed",
    )

    # -- Three open lines resolve the same way as two.
    c = one(ambiguous + [ns("3", 25, is_open=True)], vendor(58))
    check(c.status == mt.STATUS_NEEDS_RESOLUTION and len(c.candidate_lines) == 3,
          "three open lines -> NEEDS_RESOLUTION with all three surfaced",
          f"{c.status}/{len(c.candidate_lines)}")

    # -- No open line at all. Did not occur on the live population; must be a
    #    defined outcome rather than a crash.
    c = one([ns("1", 50, is_open=False, recv=50.0, billed=50.0),
             ns("2", 200, is_open=False, recv=200.0, billed=200.0)], vendor(58))
    check(c.status == mt.STATUS_NEEDS_ATTENTION,
          "no open line -> NEEDS_ATTENTION, not NEEDS_RESOLUTION", c.status)
    check("none is open" in c.attention_reason, "reason says none is open", c.attention_reason[:80])
    check(len(c.candidate_lines) == 2, "and both lines are still surfaced for the reviewer")
    check(c.line_id is None, "nothing was targeted")

    # -- A single line that is not open is not writable either. The open filter is
    #    the general rule, not a duplicates-only special case.
    c = one([ns("7", 40, is_open=False, recv=40.0, billed=40.0)], vendor(58))
    check(c.status == mt.STATUS_NEEDS_ATTENTION,
          "a lone NOT-OPEN line flags instead of proposing", c.status)
    check("none is open" in c.attention_reason, "with the same reason", c.attention_reason[:80])

    # -- Deliberately closed lines keep their own wording and their own guard.
    #    NetSuite reports a closed line as isOpen=False, so this arrives through
    #    the no-open branch, but the outcome must not degrade.
    c = one([ns("7", 40, is_open=False, closed=True)], vendor(58))
    check(c.line_closed, "a closed line still sets line_closed")
    check("closed in NetSuite" in c.attention_reason,
          "and still reads as closed, not merely 'not open'", c.attention_reason[:80])
    expect_raises(
        mt.LineClosed,
        lambda: c.to_netsuite_fields(include_dates=False),
        "and to_netsuite_fields still raises LineClosed",
    )

    # -- The ordinary case is untouched: one matching line, no duplicates.
    c = one([ns("5", 12)], vendor(9))
    check(c.status == mt.STATUS_PENDING_REVIEW and c.line_id == "5",
          "a single open match still proposes normally", f"{c.status}/{c.line_id}")
    check(c.candidate_lines == [], "and carries no candidate payload")

    # -- Genuinely no match still reads as 'no line matches', not 'none is open'.
    c = one([ns("5", 12, color="BLK")], vendor(9))
    check(c.status == mt.STATUS_NEEDS_ATTENTION and "no NetSuite line" in c.attention_reason,
          "an unmatched vendor line keeps its own distinct reason", c.attention_reason[:70])

    # -- Duplicate NetSuite lines must not confuse the reporting helper either:
    #    one vendor line covers BOTH twins, so neither is reported unshipped.
    left = mt.unmatched_netsuite_lines([vendor(58)], ambiguous + [ns("9", 4, size="XL")])
    check([l.line_id for l in left] == ["9"],
          "unmatched_netsuite_lines treats both twins as shipped, not just the first",
          str([l.line_id for l in left]))

    # -- And the extraction side is UNCHANGED: two rows for the same key in one
    #    document still sum. Same symptom, different problem, opposite answer.
    from extraction_schema import aggregate_lines

    rows = [{"po_number": "1649", "style_number": "A320001", "color": "WHT", "size": "ALL",
             "quantity": 50, "confidence": "high", "note": "", "source_hint": "P1!R5"},
            {"po_number": "1649", "style_number": "A320001", "color": "WHT", "size": "ALL",
             "quantity": 8, "confidence": "high", "note": "", "source_hint": "P1!R9"}]
    agg, _ = aggregate_lines(rows)
    check(len(agg) == 1 and agg[0]["quantity"] == 58,
          "extraction-side duplicates still sum to 58 (change 2 intact)",
          str([(l["quantity"]) for l in agg]))


def test_scope_boundaries(tmp: Path) -> None:
    section("scope boundaries (things this tool must never do)")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    ns = POLine(
        line_id="5", item="A320001 : A320001-WHT-ALL", style_number="A320001",
        vendor_name=None, color="WHT", size="ALL", quantity=12, units="Ea",
        expected_receipt_date=dt.date(2026, 9, 1), override_expected_receipt=False,
        updated_receipt_date=None, is_open=True, quantity_received=0.0,
        quantity_billed=0.0, rate=12.5,
    )
    client = NetSuiteClient(mock_data={"1649": [ns]})
    vendor = {"po_number": "1649", "style_number": "A320001", "color": "WHT",
              "size": "ALL", "quantity": 20, "confidence": "high", "note": ""}
    c = mt.build_proposed_changes([vendor], client)[0]

    # 1. The tool NEVER creates PO lines -- it only updates existing ones. A
    #    vendor line with no NetSuite line is a flag, never an insert.
    check(not [m for m in dir(client) if m.startswith(("create", "add_", "insert"))],
          "the client exposes no create/insert method at all",
          str([m for m in dir(client) if m.startswith(("create", "add_", "insert"))]))
    orphan = mt.build_proposed_changes(
        [{**vendor, "style_number": "W999999"}], client)[0]
    check(orphan.status == mt.STATUS_NEEDS_ATTENTION and orphan.line_id is None,
          "a vendor line with no NetSuite line flags -- it does not become a new line",
          orphan.status)

    # 2. The tool NEVER derives a date from a vendor document. The vendor's ETD/ETA
    #    are carried as reference text only; a write that would include a date no
    #    human confirmed raises rather than quietly omitting it.
    dated = mt.build_proposed_changes([vendor], client, eta="2026/7/1 08:00", etd="2026/6/20 08:00")[0]
    check(dated.vendor_eta == "2026-07-01" and dated.vendor_etd == "2026-06-20",
          "vendor dates are carried for display", f"{dated.vendor_etd}/{dated.vendor_eta}")
    check(not any("proposed" in f and "date" in f for f in mt.asdict(dated)),
          "ProposedChange has no proposed_*_date field for a date to leak into",
          str([f for f in mt.asdict(dated) if "date" in f]))
    expect_raises(
        mt.DateNotConfirmed,
        lambda: dated.to_netsuite_fields(include_dates=True),
        "a date write with only vendor dates available RAISES DateNotConfirmed",
    )
    check(dated.to_netsuite_fields(include_dates=False) == {"quantity": 20},
          "the quantity-only write carries no date field of any kind",
          str(dated.to_netsuite_fields(include_dates=False)))

    # Once Paula confirms a date, the override mechanism is what gets written:
    # custcol_override_expected_receipt = True plus custcol_sd_updatedreceiptdate.
    dated.confirm_receipt_date("2026-09-10")
    fields = dated.to_netsuite_fields(include_dates=True)
    check(fields.get(mt.NS_OVERRIDE_EXPECTED_RECEIPT) is True,
          "confirmed date sets custcol_override_expected_receipt = True")
    check(fields.get(mt.NS_UPDATED_RECEIPT_DATE) == "2026-09-10",
          "and writes the confirmed date to custcol_sd_updatedreceiptdate",
          str(fields.get(mt.NS_UPDATED_RECEIPT_DATE)))
    # OPEN QUESTION, recorded here rather than assumed either way: expectedReceiptDate
    # is currently written too, with the same confirmed value. If the override
    # mechanism alone is meant to drive it, this field should come out of the write
    # -- that is Paula's call, not a refactor. This assertion pins today's behaviour
    # so the decision is visible when it is made, not silently changed.
    check(fields.get(mt.NS_EXPECTED_RECEIPT_DATE) == "2026-09-10",
          "expectedReceiptDate ALSO carries the confirmed date (see comment: open question)",
          str(fields.get(mt.NS_EXPECTED_RECEIPT_DATE)))

    # 3. The tool NEVER touches custcol_sd_fg_excluderepspark. Paula manages it by
    #    hand. It is not read, not written, not displayed -- including on the
    #    resolution payload, where it would be the obvious thing to show.
    for label, payload in (("the staged change", mt.asdict(c)),
                           ("the write dict", dated.to_netsuite_fields(include_dates=True)),
                           ("the quantity-only write", c.to_netsuite_fields(include_dates=False))):
        check(all("repspark" not in str(k).lower() for k in payload),
              f"no repspark field in {label}", str(list(payload)[:4]))
    check("repspark" not in str(ns.raw).lower() and not hasattr(ns, "exclude_repspark"),
          "and POLine does not model it either -- the field is never read from NetSuite")


def test_unopenable_files(tmp: Path) -> None:
    section("corrupt / password-protected files flag, they don't crash the job")
    corrupt = HERE / "fixtures" / "corrupt_truncated.xlsx"
    encrypted = HERE / "fixtures" / "encrypted_password_protected.pdf"
    if not (corrupt.exists() and encrypted.exists()):
        _missing_coverage.append(
            "unopenable-file tests skipped: run `python make_test_fixtures.py` to build fixtures"
        )
        print("  [MISSING] fixtures absent -- run make_test_fixtures.py")
        return

    # 1. The openers classify each failure specifically.
    try:
        ce.open_workbook(corrupt)
        check(False, "corrupt xlsx raises DocumentUnreadable", "no error")
    except ce.DocumentUnreadable as exc:
        check(True, "corrupt xlsx raises DocumentUnreadable")
        check("zip archive" in exc.reason and "truncated" in exc.reason,
              "reason names the real cause (xlsx is a zip)", exc.reason[:70])
        check(exc.path.name == corrupt.name, "reason carries the offending file")

    try:
        with ce.open_pdf(encrypted) as pdf:
            pdf.pages[0].extract_text()
        check(False, "encrypted PDF raises DocumentUnreadable", "no error")
    except ce.DocumentUnreadable as exc:
        check(True, "encrypted PDF raises DocumentUnreadable")
        check("encrypt" in exc.reason.lower() or "password" in exc.reason.lower(),
              "reason mentions encryption/password", exc.reason[:80])

    check(isinstance(ce.DocumentUnreadable(corrupt, "x"), ce.ExtractionError),
          "DocumentUnreadable is an ExtractionError, so existing handlers still catch it")

    # 2. Higher-level readers surface it rather than leaking a zipfile error.
    expect_raises(ce.DocumentUnreadable, lambda: ce.read_workbook_grids(corrupt),
                  "read_workbook_grids raises DocumentUnreadable")
    expect_raises(ce.DocumentUnreadable, lambda: ce.read_pdf_layouts(encrypted),
                  "read_pdf_layouts raises DocumentUnreadable")
    expect_raises(ce.DocumentUnreadable, lambda: dp.extract_pdf_text(encrypted),
                  "extract_pdf_text raises DocumentUnreadable")

    # The Inprotex sniffer must answer "no" with a specific reason, not explode.
    matched, reason = dp.looks_like_inprotex(corrupt)
    check(not matched, "format sniff on a corrupt file returns a routing answer", reason[:60])
    check("could not open" in reason, "with the specific open failure named")

    # 3. THE POINT: a bad attachment must not take the batch down with it.
    good = tmp / "inprotex_like.xlsx"
    write_inprotex_like(good)
    sources, warnings = dp.build_source_documents([corrupt, good, encrypted])
    check(bool(sources), "the good document was still rendered", f"{len(sources)} source(s)")
    check(all("inprotex_like" in s.label for s in sources), "and only the good one")
    check(sum("COULD NOT OPEN" in w for w in warnings) == 2,
          "both bad attachments flagged", str([w[:44] for w in warnings]))
    check(any(corrupt.name in w for w in warnings), "corrupt file named in a warning")
    check(any(encrypted.name in w for w in warnings), "encrypted file named in a warning")
    check(any("still processed" in w for w in warnings), "warning says the batch continued")

    # 4. Triage reports it as its own condition, distinct from "no size data".
    import attachment_classifier as ac

    result = ac.classify_attachments([corrupt, encrypted], use_content_check=False)
    check(result.needs_manual_entry, "an email of only unopenable files -> manual entry")
    check(len(result.excluded) == 2, "both excluded", str(len(result.excluded)))
    for item in result.excluded:
        check(item.unreadable_reason is not None, f"{item.path.name[:24]} marked unreadable")
        check(item.excluded_reason.startswith("could not open"),
              "excluded for being unopenable, not for lacking sizes", item.excluded_reason[:44])
    check(any("resend" in w for w in result.warnings), "suggests asking the vendor to resend")

    check(ac.open_failure_reason(good) is None, "a good file reports no open failure")
    check(ac.open_failure_reason(corrupt) is not None, "a corrupt file reports one")


def test_size_value_space_failsafe(tmp: Path) -> None:
    section("unrecognized size labels fail safe (see RUNBOOK §6 item 4)")
    import datetime as dt

    import matcher as mt
    from netsuite_client import NetSuiteClient, POLine

    # The catalog-wide size value space could NOT be established -- every
    # introspection path is blocked for the least-privilege role (SuiteQL,
    # collection queries, and the size custom list are all denied). So rather
    # than guessing at numeric-size aliases, assert the fail-safe: an
    # unrecognized size must NEVER silently mis-match.
    client = NetSuiteClient(
        mock_data={
            "1721": [
                POLine(
                    line_id="10", item="W520005 : W520005-Pants-BLACK-M", style_number="W520005",
                    vendor_name=None, color="BLACK", size="M", quantity=10, units="Ea",
                    expected_receipt_date=dt.date(2026, 9, 1), override_expected_receipt=False,
                    updated_receipt_date=None,
                )
            ]
        }
    )
    for odd_size in ("32", "W32L34", "34x30", "One Size", ""):
        change = mt.build_proposed_changes(
            [{"po_number": "1721", "style_number": "W520005", "color": "BLACK",
              "size": odd_size, "quantity": 5, "confidence": "high", "note": ""}],
            client,
        )[0]
        check(
            change.status == mt.STATUS_NEEDS_ATTENTION,
            f"size {odd_size!r} flags rather than mis-matching the 'M' line",
            change.status,
        )
        check(change.proposed_quantity == 5 and change.current_quantity is None,
              f"size {odd_size!r} produced no silent quantity overwrite")

    # And a known alias still matches, so the fail-safe isn't just rejecting everything.
    ok = mt.build_proposed_changes(
        [{"po_number": "1721", "style_number": "W520005", "color": "BLACK",
          "size": "M", "quantity": 5, "confidence": "high", "note": ""}],
        client,
    )[0]
    check(ok.status == mt.STATUS_PENDING_REVIEW, "a recognized size still matches normally", ok.status)


def test_no_packing_sheet_becomes_manual_entry(tmp: Path) -> None:
    section("no-packing-sheet workbook -> needs_manual_entry, diagnostics intact")
    none_ = HERE / "fixtures" / "no_packing_sheet.xlsx"
    if not none_.exists():
        _missing_coverage.append(
            "manual-entry conversion test skipped: run `python make_test_fixtures.py`"
        )
        print("  [MISSING] fixture absent -- run make_test_fixtures.py")
        return

    import attachment_classifier as ac

    def verdicts(*specs):
        return response(
            ac._ContentVerdicts(
                verdicts=[
                    ac._ContentVerdict(doc_type=dt, has_size_breakdown=sz, reason=why)
                    for dt, sz, why in specs
                ]
            )
        )

    # Triage picks the workbook (it looks like a packing list from the outside),
    # then sheet-level classification finds no packing sheet inside it.
    saved = ac.classify_attachments
    try:
        def fake_triage(paths, extractor=None, use_content_check=True):
            res = ac.ClassificationResult()
            res.selected = [
                ac.AttachmentClassification(
                    path=Path(p), doc_type=ac.DocType.PACKING_LIST, has_size_breakdown=True,
                    reason="filename+preview suggested a packing list", method="filename+content",
                )
                for p in paths
            ]
            return res

        ac.classify_attachments = fake_triage
        client, parse = fake_client([
            verdicts(
                ("commercial_invoice", False, "COMMERCIAL INVOICE header, no size columns"),
                ("other", False, "totals only, no line data"),
            )
        ])
        result = dp.parse_shipment_email([none_], extractor=ce.ClaudeExtractor(client=client))
    finally:
        ac.classify_attachments = saved

    # The point: an outcome, not an exception.
    check(result.needs_manual_entry, "returns needs_manual_entry instead of raising")
    check(result.lines == [], "no lines invented", str(len(result.lines)))
    check(result.needs_review, "and needs_review is set")
    check("no packing sheet" in result.parser, "parser records what happened", result.parser)
    check(any("MANUAL ENTRY REQUIRED" in w for w in result.warnings), "says manual entry unmistakably")
    check(any("NOT an empty shipment" in w for w in result.warnings),
          "distinguishes itself from an empty shipment")
    check(any("Do not infer" in w for w in result.warnings),
          "restates that sizes must not be inferred")

    # REQUIRED: the per-sheet diagnostic must survive, not be flattened.
    joined = " || ".join(result.warnings)
    check("'INVOICE'" in joined, "warning set names the INVOICE sheet")
    check("'SUMMARY'" in joined, "warning set names the SUMMARY sheet")
    check("commercial_invoice" in joined, "carries the predicted type for a sheet")
    check(joined.count("predicted type") >= 2, "one predicted-type entry per sheet",
          str(joined.count("predicted type")))
    check(any("full extractor diagnostic" in w for w in result.warnings),
          "the original exception message is preserved verbatim too")
    check(any(none_.name in w for w in result.warnings), "and the workbook is named")
    check(len(parse.calls) == 1, "only the classification call was spent", str(len(parse.calls)))

    # parse_packing_slip itself still raises -- the conversion belongs at the
    # email level, so the lower-level API keeps its precise failure.
    client, _ = fake_client([
        verdicts(
            ("commercial_invoice", False, "no size columns"),
            ("other", False, "totals only"),
        )
    ])
    expect_raises(
        ce.NoPackingSheetFound,
        lambda: dp.parse_packing_slip(none_, extractor=ce.ClaudeExtractor(client=client), force="claude"),
        "parse_packing_slip still raises (conversion is only at the email level)",
    )


def test_manual_entry_path(tmp: Path) -> None:
    section("shipment with no size-level source -> manual entry, never inferred")
    inprotex = tmp / "inprotex_like.xlsx"
    write_inprotex_like(inprotex)

    import attachment_classifier as ac

    # Simulate triage finding nothing usable.
    saved = ac.classify_attachments
    try:
        def fake_classify(paths, extractor=None, use_content_check=True):
            res = ac.ClassificationResult()
            res.excluded = [
                ac.AttachmentClassification(
                    path=Path(p), doc_type=ac.DocType.INSPECTION_REPORT,
                    has_size_breakdown=True, reason="has sizes but is banned",
                    method="filename",
                )
                for p in paths
            ]
            res.warnings = ["no attachment provides per-size quantities"]
            return res

        ac.classify_attachments = fake_classify
        result = dp.parse_shipment_email([inprotex], extractor=ce.ClaudeExtractor(client=fake_client([])[0]))
    finally:
        ac.classify_attachments = saved

    check(result.needs_manual_entry, "result flags manual entry")
    check(result.lines == [], "NO lines produced -- sizes are not inferred from a banned document")
    check(result.needs_review, "and it needs review")
    check(any("MANUAL ENTRY REQUIRED" in w for w in result.warnings), "says so unmistakably")
    check(result.parser == "attachment-triage-only", "parser records that nothing was parsed", result.parser)


def test_matcher_handoff(tmp: Path) -> None:
    section("handoff into matcher.py (the contract that must not drift)")
    from matcher import build_proposed_changes
    from netsuite_client import NetSuiteClient, POLine
    import datetime as dt

    inprotex = tmp / "inprotex_like.xlsx"
    write_inprotex_like(inprotex)
    result = dp.parse_packing_slip(inprotex)

    for key in ("po_number", "style_number", "color", "size", "quantity"):
        check(all(key in ln for ln in result.lines), f"every line carries '{key}' for the matcher")

    # NetSuite holds 2X (canonical); the vendor sheet said XXL. The extractor must
    # NOT have normalized it -- matcher.py's SIZE_ALIASES does, and this proves the
    # split works rather than both sides guessing.
    mock_state = {
        "1662": [
            POLine(
                line_id="18", item="M120246 : M120246-Waterman Polo-TID-2X",
                style_number="M120246", vendor_name=None, color="TID", size="2X",
                quantity=2, units="Ea", expected_receipt_date=dt.date(2026, 7, 15),
                override_expected_receipt=False, updated_receipt_date=None,
            )
        ]
    }
    changes = build_proposed_changes(result.lines, NetSuiteClient(mock_data=mock_state), eta="2026/6/27 16:45")
    xxl = [c for c in changes if c.size == "XXL"]
    check(len(xxl) == 1, "the XXL line reached the matcher")
    check(xxl[0].status == "PENDING_REVIEW", "XXL matched NetSuite's 2X line via SIZE_ALIASES", xxl[0].status)
    check(xxl[0].current_quantity == 2 and xxl[0].proposed_quantity == 4, "diff computed", "2 -> 4")
    unmatched = [c for c in changes if c.status == "NEEDS_ATTENTION"]
    check(len(unmatched) == 2, "S and M have no mock NetSuite line -> NEEDS_ATTENTION, not dropped", str(len(unmatched)))


def test_schema_guarantees() -> None:
    section("schema guarantees")
    from pydantic import ValidationError

    check(ExtractedLine.model_json_schema()["required"], "all line fields are required (structured outputs need it)")
    try:
        ExtractedLine(po_number="1", style_number="s", color="c", size="S", quantity=1,
                      confidence="pretty sure", note="", source_hint="")
        check(False, "an invalid confidence value is rejected", "no error raised")
    except ValidationError:
        check(True, "an invalid confidence value is rejected")

    d = __import__("extraction_schema").line_to_dict(line(size=" XXL ", note=" x "))
    check(d["size"] == "XXL" and d["note"] == "x", "line_to_dict strips whitespace")
    check(set(d) >= {"po_number", "style_number", "color", "size", "quantity"}, "matcher keys present")
    dd = __import__("extraction_schema").deterministic_line_to_dict(
        {"po_number": "1662", "style_number": "M1", "color": "TID", "size": "S", "quantity": 9.0}
    )
    check(dd["quantity"] == 9 and isinstance(dd["quantity"], int), "deterministic quantity coerced to int")
    check(dd["confidence"] == "high", "deterministic lines are high confidence")


# ---------------------------------------------------------------------------
# B. Real-sample tests
# ---------------------------------------------------------------------------


def test_real_samples() -> None:
    section("REAL sample documents")
    if not REAL_XLSX.exists():
        _missing_coverage.append(
            f"Packing slip not found: {REAL_XLSX.name}\n"
            "      -> the deterministic Inprotex parser is NOT re-verified by this suite, and\n"
            "         the Claude extractor has never seen a real vendor document."
        )
        print(f"  [MISSING] real packing slip absent: {REAL_XLSX.name}")
    else:
        matched, reason = dp.looks_like_inprotex(REAL_XLSX)
        check(matched, "real packing slip sniffs as the Inprotex layout", reason)
        result = dp.parse_packing_slip(REAL_XLSX)
        check(result.parser == "inprotex-deterministic", "used the free validated parser", result.parser)
        check(len(result.lines) == 77, "77 lines, matching the hand-verified count", str(len(result.lines)))
        check(not result.needs_review, "clean parse of the known-good file", result.review_summary())
        check(
            all(ln["po_number"] and ln["style_number"] and ln["color"] and ln["size"] for ln in result.lines),
            "every line fully populated",
        )

    if not REAL_PDF.exists():
        _missing_coverage.append(
            f"Shipping advice not found: {REAL_PDF.name}\n"
            "      -> the deterministic ETD/ETA/HAWB regex path is NOT verified by this suite."
        )
        print(f"  [MISSING] real shipping advice absent: {REAL_PDF.name}")
    else:
        info, parser, warnings = dp.parse_shipping_advice(REAL_PDF)
        check(parser == "regex-deterministic", "real advice parsed by the free regex path", parser)
        check(info.get("hawb") == "6128990769", "HAWB matches the filename", str(info.get("hawb")))
        check(info.get("mawb") == "695-59832010", "MAWB extracted", str(info.get("mawb")))
        check(info.get("invoice_no") == "SD-219", "invoice number extracted", str(info.get("invoice_no")))

        # Confirmed against the PDF's own header row ("PORT OF ORIGN ETD PORT OF
        # DEST ETA"): 19:40 is the departure, 16:45 the arrival. The ETA being
        # chronologically earlier is correct for an eastbound trans-Pacific leg,
        # and is exactly why these are matched by column label, not date order.
        check(info.get("etd") == "2026/6/27 19:40", "ETD read from the ETD column", str(info.get("etd")))
        check(info.get("eta") == "2026/6/27 16:45", "ETA read from the ETA column", str(info.get("eta")))
        check(warnings == [], "no warnings -- labels resolved unambiguously", str(warnings))

        # The ETA is what the diff engine proposes as a receipt date, so its
        # round-trip through the matcher's date parser matters.
        from matcher import _parse_eta_to_date
        import datetime as _dt

        check(
            _parse_eta_to_date(info["eta"]) == _dt.date(2026, 6, 27),
            "ETA parses to the date §6.1 documents for this shipment",
            str(_parse_eta_to_date(info["eta"])),
        )


# ---------------------------------------------------------------------------
# C. Live Claude test (opt-in)
# ---------------------------------------------------------------------------


def test_live(tmp: Path) -> None:
    section("LIVE Claude extraction (--live)")
    if not ce.credentials_available():
        _missing_coverage.append(
            "Live Claude path not exercised: no Anthropic credential found.\n"
            "      -> set ANTHROPIC_API_KEY or run `ant auth login`, then re-run with --live."
        )
        print("  [MISSING] no Anthropic credential; skipping the live call")
        return

    path = tmp / "unknown_vendor.xlsx"
    write_unknown_vendor(path)
    try:
        result = dp.parse_packing_slip(path, force="claude")
    except Exception:  # noqa: BLE001
        print()
        traceback.print_exc()
        check(False, "live extraction of a synthetic unknown-vendor sheet")
        return

    check(len(result.lines) == 2, "extracted both rows", str(len(result.lines)))
    sizes = {ln["size"] for ln in result.lines}
    check(sizes == {"XS", "S"}, "sizes read verbatim", str(sizes))
    check(all(ln["po_number"] for ln in result.lines), "PO number read from an unfamiliar column name")
    check(result.usage.get("input_tokens", 0) > 0, "token usage reported", str(result.usage))
    print(f"    layout read as: {result.document_summary}")


def test_live_legendz(tmp: Path) -> None:
    section("LIVE — Legendz (2nd real vendor: sizes as columns, subtotal rows)")
    if not ce.credentials_available():
        print("  [MISSING] no Anthropic credential; skipping")
        return
    if not LEGENDZ_XLSX.exists():
        _missing_coverage.append(f"Legendz generalization test skipped: {LEGENDZ_XLSX.name} absent")
        print(f"  [MISSING] {LEGENDZ_XLSX.name} absent")
        return

    matched, reason = dp.looks_like_inprotex(LEGENDZ_XLSX)
    check(not matched, "not misrouted to the Inprotex parser", reason)

    result = dp.parse_packing_slip(LEGENDZ_XLSX)
    check(result.parser == "claude-assisted", "routed to the Claude extractor", result.parser)

    got = {
        (l["po_number"], l["style_number"], l["color"], l["size"]): l["quantity"] for l in result.lines
    }
    check(len(result.lines) == len(LEGENDZ_EXPECTED), f"{len(LEGENDZ_EXPECTED)} lines", str(len(result.lines)))
    check(got == LEGENDZ_EXPECTED, "every PO/style/colour/size/quantity correct",
          f"wrong: { {k: (got.get(k), v) for k, v in LEGENDZ_EXPECTED.items() if got.get(k) != v} or 'none' }")

    # The subtotal trap: rows 11/14/18 and the grand total must not become lines.
    total = sum(got.values())
    check(total == 1049, "sum equals the sheet's GRAND TOTAL -- no double counting", str(total))
    check(
        not any(q in (573, 328, 1049) for q in got.values()),
        "no line carries a printed subtotal or grand-total figure",
    )
    check(
        all(k[0] == "1657" for k in got) and {k[1] for k in got} == {"M630018", "M680009"},
        "PO/style split correctly on the full-width comma in 'PO#1657，M630018'",
        str({k[1] for k in got}),
    )
    check({k[2] for k in got} == {"DFK", "MLT", "DKF"}, "colour codes verbatim, incl. the DFK/DKF pair", str({k[2] for k in got}))
    check("2XL" in {k[3] for k in got}, "2XL kept verbatim, not converted")

    ship, _ = dp.parse_shipping_info_from_documents([LEGENDZ_XLSX])
    check(ship.get("etd") == "2026/8/5", "ETD from the free-text cell", str(ship.get("etd")))
    check(ship.get("eta") == "2026/8/16", "ETA from the free-text cell (no labelled table exists)", str(ship.get("eta")))
    print(f"    tokens: {result.usage}")


def _aggregate(lines: list[dict]) -> dict:
    out: dict[tuple, int] = {}
    for ln in lines:
        key = (ln["po_number"], ln["style_number"], ln["color"], ln["size"])
        out[key] = out.get(key, 0) + ln["quantity"]
    return out


def test_live_symmetry(tmp: Path) -> None:
    section("LIVE — Symmetry (3rd real vendor: PDF packing list, sparse size columns)")
    if not ce.credentials_available():
        print("  [MISSING] no Anthropic credential; skipping")
        return
    if not (SYMMETRY_COVERING.exists() and SYMMETRY_DETAIL.exists()):
        _missing_coverage.append("Symmetry test skipped: real packing lists absent")
        print("  [MISSING] Symmetry packing lists absent")
        return

    # PRIMARY: the style/colour/size rollup.
    covering = dp.parse_packing_slip(SYMMETRY_COVERING)
    cov = _aggregate(covering.lines)
    check(covering.parser == "claude-assisted", "PDF packing list routed to the Claude path", covering.parser)
    check(len(cov) == len(SYMMETRY_EXPECTED), f"{len(SYMMETRY_EXPECTED)} size-level keys", str(len(cov)))
    check(cov == SYMMETRY_EXPECTED, "every PO/style/colour/size/quantity correct",
          f"wrong: { {k: (cov.get(k), v) for k, v in SYMMETRY_EXPECTED.items() if cov.get(k) != v} or 'none' }")
    check(sum(cov.values()) == SYMMETRY_GRAND_TOTAL, "sum equals the printed G.TOTAL", str(sum(cov.values())))

    # Sparse size columns — the alignment trap on this document.
    check(("1720", "M650022", "NEW INDIGO", "XS") not in cov, "M650022 correctly has NO XS")
    check(cov.get(("1720", "M650022", "NEW INDIGO", "S")) == 22, "its first figure is S=22, not XS=22")
    check(
        all(("1721", "W520005", "COCONUT", s) not in cov for s in ("XL", "2XL", "3XL")),
        "W520005 COCONUT correctly stops at L",
    )
    check(cov.get(("1720", "M650022", "NEW INDIGO", "3XL")) == 4, "3XL=4 read from the last column")

    # CROSS-CHECK: the carton-by-carton file must agree with the rollup. It spans
    # two pages with headers only on page 1, which is why all pages go in one call.
    detail = dp.parse_packing_slip(SYMMETRY_DETAIL)
    det = _aggregate(detail.lines)
    check(det == cov, "carton detail agrees exactly with the rollup",
          f"diffs: { {k: (cov.get(k), det.get(k)) for k in set(cov) | set(det) if cov.get(k) != det.get(k)} or 'none' }")
    check(sum(det.values()) == SYMMETRY_GRAND_TOTAL, "detail also sums to the G.TOTAL", str(sum(det.values())))

    # Sanity only. Inspection reports are NOT a data source (Paula's ruling); this
    # just confirms the packing lists are internally consistent with numbers we
    # independently validated earlier.
    insp = {
        ("1721", "W600001", "BLACK", s): q
        for s, q in (("XS", 19), ("S", 71), ("M", 110), ("L", 82), ("XL", 39))
    }
    check(all(cov.get(k) == v for k, v in insp.items()),
          "W600001 BLACK consistent with previously validated figures (sanity check only)")
    print(f"    tokens: covering={covering.usage} detail={detail.usage}")


def test_live_attachment_triage(tmp: Path) -> None:
    section("LIVE — attachment triage on the real six-attachment Symmetry email")
    if not ce.credentials_available():
        print("  [MISSING] no Anthropic credential; skipping")
        return
    present = [
        p for p in (SYMMETRY_INVOICE, SYMMETRY_COVERING, SYMMETRY_DETAIL,
                    SYMMETRY_PAYMENT_REQUEST, SYMMETRY_INSPECTION, SYMMETRY_INSPECTION_2)
        if p.exists()
    ]
    if len(present) < 5:
        _missing_coverage.append("attachment triage test skipped: not all Symmetry attachments present")
        print(f"  [MISSING] only {len(present)} attachments present")
        return

    import attachment_classifier as ac

    result = ac.classify_attachments(present)
    selected = {c.path.name for c in result.selected}
    excluded = {c.path.name: c for c in result.excluded}

    check(SYMMETRY_COVERING.name in selected, "the rollup packing list is selected")
    check(SYMMETRY_DETAIL.name in selected, "the carton detail packing list is selected")
    check(result.primary.path.name == SYMMETRY_COVERING.name,
          "the rollup is chosen as PRIMARY (already in the target shape)", result.primary.path.name)

    # THE TRAP: filename says PACKING LIST, content is a customs invoice with no
    # sizes. Parsing this is what produced the wrong "Symmetry has no size
    # breakdown" conclusion, so rejecting it is the regression that matters most.
    check(SYMMETRY_INVOICE.name in excluded, "the customs invoice is REJECTED despite its filename")
    if SYMMETRY_INVOICE.name in excluded:
        item = excluded[SYMMETRY_INVOICE.name]
        check(item.doc_type == ac.DocType.COMMERCIAL_INVOICE,
              "classified as a commercial invoice by content, not by name", item.doc_type.value)

    for insp in (SYMMETRY_INSPECTION, SYMMETRY_INSPECTION_2):
        if insp.exists():
            check(insp.name in excluded, f"inspection report excluded: {insp.name[:34]}")
            check("Paula" in excluded[insp.name].excluded_reason,
                  "excluded on the standing ruling, not on content")
    if SYMMETRY_PAYMENT_REQUEST.exists():
        check(SYMMETRY_PAYMENT_REQUEST.name in excluded, "payment request excluded")
    check(not result.needs_manual_entry, "a usable packing list was found, so no manual entry needed")

    # Inprotex's workbook also says "Invoice" in its name and IS the real
    # size-level source — the opposite direction of the same trap. Its size header
    # sits far below the top of the PACKING sheet, so the preview must seek it out.
    if REAL_XLSX.exists():
        inpro = ac.classify_attachments([REAL_XLSX])
        check(bool(inpro.selected), "Inprotex 'Invoice_Packing.xlsx' selected (opposite filename trap)")
        if inpro.selected:
            check(inpro.selected[0].has_size_breakdown,
                  "its size header was found below the letterhead rows",
                  inpro.selected[0].reason[:80])


# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--live", action="store_true", help="also make a real Anthropic API call")
    args = ap.parse_args()

    print("=" * 78)
    print("PARSING LAYER TESTS")
    print("=" * 78)

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        offline = [
            test_grid_rendering, test_windowing, test_extractor_call_shape,
            test_failure_modes_raise, test_low_confidence_flagging, test_routing,
            test_deterministic_validation, test_shipping_advice_routing,
            test_shipping_date_label_anchoring, test_pdf_layout_rendering,
            test_multidoc_call_shape, test_attachment_classifier_offline,
            test_canonical_form, test_size_header_canonical,
            test_verbatim_source_preserved,
            test_matcher_canonical_both_sides,
            test_line_aggregation, test_aggregation_wired_into_parsers,
            test_matcher_paula_rulings, test_sheet_selection, test_closed_po_line,
            test_duplicate_key_resolution, test_scope_boundaries,
            test_unopenable_files, test_size_value_space_failsafe,
            test_no_packing_sheet_becomes_manual_entry,
            test_manual_entry_path, test_matcher_handoff,
        ]
        for fn in offline:
            try:
                fn(tmp)
            except Exception:  # noqa: BLE001
                print()
                traceback.print_exc()
                _results.append((False, f"{fn.__name__} crashed", ""))

        try:
            test_schema_guarantees()
        except Exception:  # noqa: BLE001
            print()
            traceback.print_exc()
            _results.append((False, "test_schema_guarantees crashed", ""))

        try:
            test_real_samples()
        except Exception:  # noqa: BLE001
            print()
            traceback.print_exc()
            _results.append((False, "test_real_samples crashed", ""))

        if args.live:
            for fn in (test_live, test_live_legendz, test_live_symmetry, test_live_attachment_triage):
                try:
                    fn(tmp)
                except Exception:  # noqa: BLE001
                    print()
                    traceback.print_exc()
                    _results.append((False, f"{fn.__name__} crashed", ""))
        else:
            _missing_coverage.append(
                "Live Claude path not exercised (no --live flag).\n"
                "      -> the extractor's prompt is only tested against a mock in this run.\n"
                "         Run with --live to check it against all three real vendors."
            )

    passed = sum(1 for ok, _, _ in _results if ok)
    failed = [name for ok, name, _ in _results if not ok]

    print()
    print("=" * 78)
    print(f"{passed}/{len(_results)} checks passed")
    if failed:
        print()
        for name in failed:
            print(f"  FAILED: {name}")

    if _missing_coverage:
        print()
        print("!" * 78)
        print("MISSING TEST COVERAGE -- a green run above does NOT mean this is proven")
        print("!" * 78)
        for gap in _missing_coverage:
            print(f"  - {gap}")
        print()
        print("  The Claude-assisted extractor is the PRIMARY parsing path for this project")
        print("  and has not been run against a single real vendor document. Getting a second")
        print("  and third real vendor file is the highest-value next step (build plan, Risks).")
    print("=" * 78)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
